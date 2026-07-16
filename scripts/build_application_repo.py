from __future__ import annotations

import json
import re
from pathlib import Path

from openpyxl import load_workbook

from active_school_packets import ACTIVE_SCHOOL_NAMES, NEW_SCHOOLS

ROOT = Path(__file__).resolve().parents[1]
SCHOOL_LIST_XLSX = ROOT / "med_school_list_md.xlsx"
AMCAS_GUIDE = Path("/Users/jinko/Downloads/aamc-2027-amcas-applicant-guide_final_0.pdf")
BUILD_DATE = "2026-05-30"


PRIMARY_REQUIREMENTS = [
    "AMCAS primary application",
    "Official transcripts sent to AMCAS",
    "Letters of evaluation assigned through AMCAS",
    "MCAT score on file",
    "School-specific secondary application when invited",
]


AMCAS_NOTES = {
    "personal_statement": "AMCAS personal comments essay: 5,300 characters including spaces.",
    "verification": "AMCAS verification may take six to eight weeks during peak season.",
    "transcripts": "Regular MD applicants must have official transcripts received within 14 calendar days of each school's AMCAS deadline.",
    "letters": "Letters are not required for AMCAS verification, but they are transmitted to schools as AMCAS receives them.",
    "activities": "AMCAS allows a maximum of 15 Work/Activities entries.",
}


SHARED_THEMES = {
    "why-school": {
        "title": "Why This School Backbone",
        "body": """# Why This School Backbone

## Core Claim
Write the one-sentence reason this school belongs on your final list.

## School-Specific Evidence
- Mission fit
- Curriculum structure
- Clinical environment
- Research, service, or community programs
- Geographic or population fit

## Your Match
- What you have already done that maps to the school
- What you want to grow into at the school
- Why this school is a better fit than a generic peer school

## Final Pass
- Replace generic wording with named programs, clinics, tracks, or communities
- Make sure at least half the draft is school-specific
""",
    },
    "diversity-equity": {
        "title": "Diversity / Equity Backbone",
        "body": """# Diversity / Equity Backbone

## Experience
Choose one concrete experience rather than listing identities or values.

## Reflection
- What did the experience teach you?
- What bias, blind spot, or system did it reveal?
- How did it change how you show up for others?

## Forward-Looking Contribution
- What will you contribute to classmates?
- What will you contribute to patients?
- How will you keep learning?
""",
    },
    "gap-year": {
        "title": "Gap Year / Current Year Backbone",
        "body": """# Gap Year / Current Year Backbone

## Timeline
List work, service, research, coursework, and personal commitments in order.

## Purpose
Explain why this year matters for your development, not just what you are doing.

## Connection to Medicine
- Skills built
- Exposure deepened
- Perspective gained
""",
    },
    "challenge": {
        "title": "Challenge / Adversity Backbone",
        "body": """# Challenge / Adversity Backbone

## Situation
Describe the challenge clearly and briefly.

## Response
- What did you do?
- What tradeoffs or hard decisions did you make?
- Who did you rely on?

## Growth
- What changed in you?
- How will this matter in medicine?
""",
    },
    "service": {
        "title": "Service / Underserved Backbone",
        "body": """# Service / Underserved Backbone

## Community
Name the population or community served with specificity and respect.

## Role
Explain what you actually did, not just what the organization does.

## Learning
- What did the community teach you?
- What structural issues became visible?
- How did this shift your view of medicine?
""",
    },
    "additional-info": {
        "title": "Additional Information Backbone",
        "body": """# Additional Information Backbone

Use this only if it adds real signal.

## Good Uses
- Important context not shown elsewhere
- Meaningful updates
- Clarifying unusual application features

## Avoid
- Repeating your personal statement
- Defensive explanations without insight
- Generic praise for the school
""",
    },
    "academic-context": {
        "title": "Academic Context Backbone",
        "body": """# Academic Context Backbone

## Context
State the issue plainly: low grades, withdrawals, extended timeline, or a break.

## Accountability
Own what is yours without overexplaining.

## Recovery
- What changed?
- What evidence shows improvement?
- What habits or systems are now different?
""",
    },
    "teamwork": {
        "title": "Teamwork Backbone",
        "body": """# Teamwork Backbone

## Team Situation
Pick a team-based example with visible interdependence.

## Your Role
- What role did you play?
- What role did others play?
- How did you adjust?

## Outcome
Show what good teamwork required from you.
""",
    },
    "mission-fit": {
        "title": "Mission Fit Backbone",
        "body": """# Mission Fit Backbone

## School Mission
Name the part of the mission that genuinely overlaps with your path.

## Evidence
Pair each mission point with a concrete experience.

## Future
Explain how that mission fit continues in medical school and beyond.
""",
    },
    "clinical": {
        "title": "Clinical Exposure Backbone",
        "body": """# Clinical Exposure Backbone

## Best Experience
Choose the single clinical experience that best changed your understanding of medicine.

## Learning
- What did you observe about patient care?
- What did you learn about physicians?
- What did you learn about yourself?
""",
    },
    "future-goals": {
        "title": "Future Goals Backbone",
        "body": """# Future Goals Backbone

## Near-Term
What do you hope to learn or become in medical school?

## Longer-Term
What kind of physician or leader are you building toward?

## Continuity
Connect your future goals to real prior experiences.
""",
    },
    "community": {
        "title": "Community Backbone",
        "body": """# Community Backbone

## Define the Community
Choose a community you actually belong to or serve in a meaningful way.

## Relationship
- What do you receive from it?
- What do you contribute to it?

## Medicine Link
How has this community shaped the physician you want to become?
""",
    },
    "values": {
        "title": "Values / Professionalism Backbone",
        "body": """# Values / Professionalism Backbone

## Value
Choose one value you can prove with behavior.

## Example
Tell the story that demonstrates the value under pressure.

## Reflection
Explain how that value will show up in medicine.
""",
    },
    "research": {
        "title": "Research Backbone",
        "body": """# Research Backbone

## Project
Briefly identify the research or scholarly work.

## Your Contribution
- What did you do?
- What did you learn?
- What challenged you?

## Relevance
Why does this matter for your medical training or future work?
""",
    },
}


SCHOOLS = [
    {
        "name": "UC Davis School of Medicine",
        "slug": "uc-davis",
        "secondary_cycle": "2024 archive on Admit.org",
        "secondary_deadline": "28 days after receiving the secondary",
        "official_url": "https://health.ucdavis.edu/mdprogram/admissions/",
        "prompt_source": "https://med.admit.org/secondary-essays/university-of-california-davis-school-of-medicine",
        "notes": "Regional connection to Northern or Central California matters enough to have its own short-answer prompt.",
        "prompts": [
            {
                "title": "Application Concerns",
                "limit": "500 characters",
                "themes": ["academic-context"],
                "text": "Discuss any elements of your application that you feel might be concerning to the Admissions Committee (This could include explanation of metric trends, institutional actions, legal violations, etc.) Please enter N/A if no concerns to report.",
            },
            {
                "title": "Regional Connection",
                "limit": "500 characters",
                "themes": ["community"],
                "text": "Do you have a connection to Northern or Central California?",
            },
            {
                "title": "UC Davis-Relevant Activities",
                "limit": "1500 characters each for up to 3 activities",
                "themes": ["service", "clinical", "mission-fit"],
                "text": "Please list the activities you would like to be considered in your application to UC Davis. The secondary activities may be used to highlight your experiences specifically relevant for UC Davis. Please only list three. Briefly describe each activity and its significance. How did it prepare you for a medical career?",
            },
        ],
    },
    {
        "name": "University of Rochester School of Medicine and Dentistry",
        "slug": "rochester",
        "secondary_cycle": "2024 archive on Admit.org",
        "secondary_deadline": None,
        "official_url": "https://www.urmc.rochester.edu/education/md/admissions",
        "prompt_source": "https://med.admit.org/secondary-essays/university-of-rochester-school-of-medicine-and-dentistry",
        "notes": "Only one prompt in the latest Admit.org archive, centered on community.",
        "prompts": [
            {
                "title": "Community",
                "limit": "200 words",
                "themes": ["community"],
                "text": "Involvement in community is one of the pillars of the University of Rochester School of Medicine education. Tell us about a community you identify with and how you are involved with it.",
            }
        ],
    },
    {
        "name": "Virginia Tech Carilion School of Medicine",
        "slug": "virginia-tech-carilion",
        "secondary_cycle": "2026-2027 live secondary portal prompts provided by applicant on 2026-07-01",
        "secondary_deadline": "December 15, 2025",
        "official_url": "https://medicine.vtc.vt.edu/admissions/application-process.html",
        "prompt_source": "https://apply-admissions.vtc.vt.edu/Security/Login.aspx",
        "notes": "Updated from the applicant's Virginia Tech Carilion secondary portal.",
        "prompts": [
            {
                "title": "Protecting a Vulnerable Trait",
                "limit": "400 words",
                "themes": ["values", "professional-identity"],
                "text": "Reflect on a trait or characteristic you are proud of that you feel is most vulnerable to being lost on your journey to becoming a physician. How do you plan to protect it?",
            },
            {
                "title": "Right Thing Versus Rules",
                "limit": "400 words",
                "themes": ["values", "challenge"],
                "text": 'Describe a time when doing the "right" thing conflicted with rules, authority, or expectations. How did you handle it?',
            },
            {
                "title": "Greatest Generational Challenge",
                "limit": "400 words",
                "themes": ["future-goals", "values"],
                "text": "Beyond the rise of AI, what do you consider the greatest challenge facing your generation, and how do you plan to address it?",
            },
        ],
    },
    {
        "name": "Vermont Larner College of Medicine",
        "slug": "vermont-larner",
        "secondary_cycle": "2026-2027 live secondary portal prompts provided by applicant on 2026-07-01",
        "secondary_deadline": "December 22, 2025",
        "official_url": "https://www.uvm.edu/larnermed/admissions",
        "prompt_source": "https://apply.med.uvm.edu/Security/Login.aspx",
        "notes": "Updated from the applicant's Vermont Larner secondary portal. Prompt 5 is optional.",
        "prompts": [
            {
                "title": "Active Learning Fit",
                "limit": "400 words",
                "themes": ["mission-fit", "teamwork"],
                "text": "The Larner College of Medicine curriculum emphasizes active in-person participation and peer co-learning. This relies on individual students preparing for, and being actively involved in, participatory class sessions. Please share your thoughts on the suitability of active learning for you. Consider the value of balancing your learning with the teaching of your classmates, potential challenges to this approach, personal contributions and growth areas, and examples of navigating a group learning process such as a flipped classroom experience.",
            },
            {
                "title": "Learning Across Difference",
                "limit": "400 words",
                "themes": ["diversity-equity"],
                "text": "The Larner College of Medicine at the University of Vermont recognizes that diversity extends beyond chosen and unchosen identities and encompasses an individual's entire experiences. Diverse environments can promote growth and provide an opportunity for reflection. Reflect on a time you learned something from someone or a group of people who are unlike yourself and how that challenged your preconceptions or biases. How will this experience influence your behavior in the future?",
            },
            {
                "title": "Professionalism",
                "limit": "300 words",
                "themes": ["values"],
                "text": "The Larner College of Medicine has core values that are reflected in our professionalism statement. How has your understanding of what professionalism means evolved over time? Reflect on a personal experience that contributed to your understanding of your own core values.",
            },
            {
                "title": "Why Larner",
                "limit": "250 words",
                "themes": ["why-school"],
                "text": "What unique opportunities at Larner College of Medicine would be meaningful to you in your education and how do these align with your larger goal of becoming a physician?",
            },
            {
                "title": "Additional Information",
                "limit": "250 words",
                "themes": ["additional-info", "gap-year"],
                "text": "Optional: Use this space to share any information, not presented elsewhere, that will clarify any aspect of your application. If you are a reapplicant, you may choose to share what you have done since your previous application.",
            },
        ],
    },
    {
        "name": "University of Colorado School of Medicine",
        "slug": "colorado",
        "secondary_cycle": "2024 archive on Admit.org",
        "secondary_deadline": None,
        "official_url": "https://medschool.cuanschutz.edu/education/md-program/admissions",
        "prompt_source": "https://med.admit.org/secondary-essays/university-of-colorado-school-of-medicine",
        "notes": "Includes branch-program prompts; keep only the ones relevant to the tracks you actually select.",
        "prompts": [
            {
                "title": "Leadership / Curiosity / Commitment",
                "limit": "500 words",
                "themes": ["mission-fit", "future-goals"],
                "text": "The pillars of our curriculum are Leadership, Curiosity, and Commitment. Tell us about how you have embodied one or more of these attributes in your path to medicine thus far. In which of these areas do you see the most opportunity for personal growth and why?",
            },
            {
                "title": "Inclusive Excellence",
                "limit": "300 words",
                "themes": ["diversity-equity"],
                "text": "Please describe how your background and/or your unique lived experiences contribute to our culture of inclusive excellence.",
            },
            {
                "title": "Fort Collins Branch",
                "limit": "1500 characters",
                "themes": ["why-school"],
                "text": "Fort Collins Branch: Please tell us why you are interested in being a part of the 4-year CUSOM at CSU (Fort Collins Branch campus).",
            },
            {
                "title": "Rural Branch",
                "limit": "1500 characters",
                "themes": ["why-school", "service"],
                "text": "Rural Branch: Why are you interested in being a rural physician? How will rural life and work fit your personal goals?",
            },
            {
                "title": "Aerospace Engineering Dual Degree",
                "limit": "No word limit",
                "themes": ["future-goals"],
                "text": "Aerospace engineering dual degree: Describe your career path up to this point and how that has led to your interest in human spaceflight.",
            },
            {
                "title": "Colorado Springs Branch",
                "limit": "1500 characters",
                "themes": ["why-school"],
                "text": "Colorado Springs Branch: Please submit a short statement regarding your interest in the Colorado Springs Branch.",
            },
        ],
    },
    {
        "name": "University of Wisconsin School of Medicine and Public Health",
        "slug": "wisconsin",
        "secondary_cycle": "2024 archive on Admit.org",
        "secondary_deadline": "November 20, 2025",
        "official_url": "https://www.med.wisc.edu/education/md/admissions/",
        "prompt_source": "https://med.admit.org/secondary-essays/university-of-wisconsin-school-of-medicine-and-public-health",
        "notes": "Very mission-forward; strong prompt overlap with competency, equity, and fit essays.",
        "prompts": [
            {
                "title": "Mission Fit",
                "limit": "500 words",
                "themes": ["mission-fit", "why-school"],
                "text": "The Admissions Committee is committed to holistic, mission-aligned evaluation of all applicants. Review our institution’s mission. Using specific examples, discuss how and why you believe you are a good fit for the University of Wisconsin School of Medicine and Public Health.",
            },
            {
                "title": "Competency 1",
                "limit": "250 words",
                "themes": ["values"],
                "text": "Choose one of the professional competencies that has been a point of pride or emphasis on your path to medical school. Using specific examples, describe how you have demonstrated it in your experiences to date.",
            },
            {
                "title": "Competency 2",
                "limit": "250 words",
                "themes": ["values"],
                "text": "Choose a different professional competency that has been another point of pride or emphasis. Using specific examples, describe how you have demonstrated this one in your experiences to date.",
            },
            {
                "title": "Health Equity Issue",
                "limit": "250 words",
                "themes": ["diversity-equity", "future-goals"],
                "text": "Choose a broader issue or policy that impacts health outcomes where you believe change is needed to advance health equity. Discuss the role you hope to play as a physician in addressing this issue.",
            },
            {
                "title": "COVID Impact",
                "limit": "250 words",
                "themes": ["additional-info"],
                "text": "In a paragraph, please share how COVID-19 impacted your application in any of the following domains: academic, volunteer, research, work, personal life.",
            },
        ],
    },
    {
        "name": "Lewis Katz School of Medicine at Temple University",
        "slug": "temple",
        "secondary_cycle": "2024 archive on Admit.org",
        "secondary_deadline": "January 15, 2026",
        "official_url": "https://medicine.temple.edu/education/md-program/admissions",
        "prompt_source": "https://med.admit.org/secondary-essays/lewis-katz-school-of-medicine-at-temple-university",
        "notes": "Campus-specific prompt means your selected program location should shape the school-fit draft.",
        "prompts": [
            {
                "title": "Why Temple",
                "limit": "2000 characters",
                "themes": ["why-school"],
                "text": "What is the nature of your interest in the Lewis Katz School of Medicine?",
            },
            {
                "title": "What Makes You Unique",
                "limit": "2000 characters",
                "themes": ["diversity-equity", "challenge"],
                "text": "Describe what makes you unique as an applicant, an obstacle that you had to overcome, or how you will contribute to the Katz community.",
            },
            {
                "title": "Campus Interest",
                "limit": "2000 characters",
                "themes": ["why-school"],
                "text": "Tell us about your special interest in the campus you selected.",
            },
            {
                "title": "Current Year Plans",
                "limit": "2000 characters",
                "themes": ["gap-year"],
                "text": "What are your plans for the current year?",
            },
            {
                "title": "COVID Impact",
                "limit": "2000 characters",
                "themes": ["additional-info"],
                "text": "Please use this space to describe to us how you were impacted academically, personally or professionally by COVID-19.",
            },
        ],
    },
    {
        "name": "Medical College of Wisconsin",
        "slug": "mcw",
        "secondary_cycle": "2026-2027 live secondary portal prompts provided by applicant on 2026-07-08",
        "secondary_deadline": "December 8, 2025",
        "official_url": "https://www.mcw.edu/education/medical-school/admissions",
        "prompt_source": "https://secondaryapplication.mcw.edu/Security/Login.aspx",
        "notes": "Updated from the applicant's MCW secondary portal. Live set has four required prompts and no reapplicant prompt in the pasted text.",
        "prompts": [
            {
                "title": "Contribution to MCW",
                "limit": "Not listed in portal text provided",
                "themes": ["mission-fit", "values"],
                "text": "At MCW, we are guided by the values of acting in caring ways, engaging in collaborative efforts, approaching our world with curiosity, advancing inclusive practices, demonstrating integrity in all that we do, and treating everyone with respect. Describe how your unique experiences, interests, and talents have embodied one or more of these values, and how they will shape the way you plan to contribute to the MCW learning community.",
            },
            {
                "title": "Why MCW",
                "limit": "Not listed in portal text provided",
                "themes": ["why-school"],
                "text": "How will MCW uniquely prepare you for your future goals?",
            },
            {
                "title": "Decision You Regret",
                "limit": "Not listed in portal text provided",
                "themes": ["challenge"],
                "text": "Recount a time when you made a decision you regret. How did it affect you, and what did you learn from the experience?",
            },
            {
                "title": "New Environment and Community Engagement",
                "limit": "Not listed in portal text provided",
                "themes": ["service", "community", "mission-fit"],
                "text": "MCW is committed to its missions of community engagement and improving health for all. MCW medical students play an important role in advancing these missions by working with and within various communities throughout Wisconsin which requires adaptability and a willingness to learn from new experiences. Describe a time where you entered a new environment or situation, how you adapted, what you learned, and how that experience helped you grow in your ability to serve others and contribute to MCW’s missions during your education.",
            },
        ],
    },
    {
        "name": "Sidney Kimmel Medical College at Thomas Jefferson University",
        "slug": "jefferson-kimmel",
        "secondary_cycle": "2026-2027 live secondary portal prompt provided by applicant on 2026-07-01",
        "secondary_deadline": "January 1, 2026",
        "official_url": "https://www.jefferson.edu/academics/colleges-schools-institutes/skmc/admissions.html",
        "prompt_source": "https://amp.jefferson.edu/Security/Login.aspx",
        "notes": "Updated from the applicant's Sidney Kimmel secondary portal. Only the diversity/equitable-care prompt was provided.",
        "prompts": [
            {
                "title": "Diversity Contribution",
                "limit": "Not listed in portal text provided",
                "themes": ["diversity-equity"],
                "text": "Sidney Kimmel Medical College defines diversity as the richness in human differences. How will your own experiences allow you to contribute to the diversity of the student body and to provide equitable and inclusive care to your future patients?",
            },
        ],
    },
    {
        "name": "Emory University School of Medicine",
        "slug": "emory",
        "secondary_cycle": "2024 archive on Admit.org",
        "secondary_deadline": "December 1, 2025",
        "official_url": "https://med.emory.edu/education/programs/md/admissions/",
        "prompt_source": "https://med.admit.org/secondary-essays/emory-university-school-of-medicine",
        "notes": "Short limits; concise storytelling matters more than background exposition.",
        "prompts": [
            {
                "title": "Curriculum / Current Year Plan",
                "limit": "200 words",
                "themes": ["gap-year"],
                "text": "List your entire curriculum plan for the academic year. If you are not in school, briefly describe your plans for the coming year.",
            },
            {
                "title": "Health-Related Experiences",
                "limit": "200 words",
                "themes": ["clinical"],
                "text": "Briefly describe your health-related experiences. Be sure to include important experiences that are in your AMCAS application, as well as any recent experiences.",
            },
            {
                "title": "Why Emory",
                "limit": "200 words",
                "themes": ["why-school"],
                "text": "Briefly describe your interest in Emory and the Emory degree program you have selected.",
            },
            {
                "title": "Service to Underserved Communities",
                "limit": "200 words",
                "themes": ["service"],
                "text": "Please describe any of your activities that have been in service to under-served communities.",
            },
            {
                "title": "Updates",
                "limit": "200 words",
                "themes": ["additional-info"],
                "text": "If you have any updates or new information to report since you have submitted your AMCAS primary application, please briefly describe below.",
            },
        ],
    },
    {
        "name": "University of Cincinnati College of Medicine",
        "slug": "cincinnati",
        "secondary_cycle": "2024 archive on Admit.org",
        "secondary_deadline": "December 1, 2025",
        "official_url": "https://med.uc.edu/education/doctor-of-medicine/admissions",
        "prompt_source": "https://med.admit.org/secondary-essays/university-of-cincinnati-college-of-medicine",
        "notes": "This is one of the more extensive secondary sets on your list.",
        "prompts": [
            {
                "title": "Traffic Violations",
                "limit": "750 characters",
                "themes": ["additional-info"],
                "text": "Have you ever been convicted of or pled guilty or no contest to any moving traffic violations? If yes, please explain fully below.",
            },
            {
                "title": "Matriculation-Year Plans",
                "limit": "1000 characters",
                "themes": ["gap-year"],
                "text": "Provide a thorough explanation of all activities planned between now and your matriculation into medical school.",
            },
            {
                "title": "Patient-Centered Care and Diversity",
                "limit": "2000 characters",
                "themes": ["diversity-equity"],
                "text": "Consider an experience in which you collaborated or were exposed to diverse backgrounds that would inform your own vision for patient-centered care. Please describe the impact the experience had on you.",
            },
            {
                "title": "Teamwork",
                "limit": "2000 characters",
                "themes": ["teamwork"],
                "text": "What does teamwork mean to you? What teamwork experiences have you had? How do you navigate your role and the role of other team members?",
            },
            {
                "title": "Application Concerns",
                "limit": "2000 characters",
                "themes": ["academic-context"],
                "text": "Discuss any elements of your application that you feel might be concerning to the Admissions Committee.",
            },
            {
                "title": "Why UCCOM",
                "limit": "2000 characters",
                "themes": ["why-school"],
                "text": "What is your specific interest in the UCCOM program? What opportunities would you take advantage of as a student here? Why?",
            },
            {
                "title": "Academic Breadth",
                "limit": "2000 characters",
                "themes": ["values"],
                "text": "How has your academic work beyond the traditional pre-medical requirements prepared you for medical school and for a career in medicine? Please highlight any service-learning activities that you have done.",
            },
            {
                "title": "Interview Preference",
                "limit": "150 characters",
                "themes": ["additional-info"],
                "text": "If invited to interview, indicate your interview preference and briefly explain it.",
            },
        ],
    },
    {
        "name": "University of Iowa Roy J. and Lucille A. Carver College of Medicine",
        "slug": "iowa-carver",
        "secondary_cycle": "2024 archive on Admit.org",
        "secondary_deadline": "December 15, 2025",
        "official_url": "https://medicine.uiowa.edu/md/admissions",
        "prompt_source": "https://med.admit.org/secondary-essays/university-of-iowa-roy-j-and-lucille-a-carver-college-of-medicine",
        "notes": "Two of these prompts are especially useful to sync with your why-school and uniqueness backbones.",
        "prompts": [
            {
                "title": "Why Carver",
                "limit": "1500 characters",
                "themes": ["why-school"],
                "text": "We understand you may be applying to multiple medical schools. Please explain your reasons for applying to the Carver College of Medicine.",
            },
            {
                "title": "Unique Characteristic",
                "limit": "1500 characters",
                "themes": ["diversity-equity", "challenge"],
                "text": "Describe a personal characteristic, challenge, or experience that makes you unique. How will this influence your contribution to a dynamic healthcare system that advocates for all peoples?",
            },
            {
                "title": "Medically Related Experiences",
                "limit": "175 characters",
                "themes": ["clinical"],
                "text": "Please list and briefly describe all medically related experiences (paid/volunteer) you have completed during the past 5 years.",
            },
            {
                "title": "Reapplicant",
                "limit": "1500 characters",
                "themes": ["additional-info"],
                "text": "If you are a reapplicant to the Carver College of Medicine, how have you strengthened your application?",
            },
            {
                "title": "If Not in Degree Program",
                "limit": "1500 characters",
                "themes": ["gap-year"],
                "text": "If you are not currently in a degree-seeking program, please indicate what you will be doing from the time you complete this secondary application to the start of medical school.",
            },
        ],
    },
    {
        "name": "Western Michigan University Homer Stryker M.D. School of Medicine",
        "slug": "western-michigan",
        "secondary_cycle": "2024 archive on Admit.org",
        "secondary_deadline": None,
        "official_url": "https://med.wmich.edu/node/132",
        "prompt_source": "https://med.admit.org/secondary-essays/western-michigan-university",
        "notes": "Strong focus on WMed-specific fit plus regional connection.",
        "prompts": [
            {
                "title": "Why WMed",
                "limit": "2000 characters",
                "themes": ["why-school"],
                "text": "Please explain both of the following: the specific reason(s) you have chosen to apply to WMed and how you will utilize the unique features of WMed's mission, vision, and curriculum to achieve your career goals.",
            },
            {
                "title": "Diversity Contribution",
                "limit": "2000 characters",
                "themes": ["diversity-equity"],
                "text": "Describe how you add to the cultural, ethnic, and socioeconomic diversity of the medical profession and what you bring to the practice of medicine.",
            },
            {
                "title": "Reapplicant",
                "limit": "2000 characters",
                "themes": ["additional-info"],
                "text": "WMed re-applicants only: Describe the changes to your application from previous cycles.",
            },
            {
                "title": "Southwest Michigan Connection",
                "limit": "2000 characters",
                "themes": ["community"],
                "text": "Please explain any connection you have to Southwest Michigan.",
            },
            {
                "title": "Additional Information",
                "limit": "2000 characters",
                "themes": ["additional-info"],
                "text": "Is there any additional information not included elsewhere in your application you would like the admissions committee to know?",
            },
        ],
    },
    {
        "name": "Geisel School of Medicine at Dartmouth",
        "slug": "geisel",
        "secondary_cycle": "2024 archive on Admit.org",
        "secondary_deadline": None,
        "official_url": "https://geiselmed.dartmouth.edu/admissions/",
        "prompt_source": "https://med.admit.org/secondary-essays/geisel-school-of-medicine-at-dartmouth",
        "notes": "The Geisel prompts let you reuse current-year plans, why-school, additional-info, and social-justice backbones.",
        "prompts": [
            {
                "title": "Current Year Plans",
                "limit": "No word limit",
                "themes": ["gap-year"],
                "text": "Please indicate your plans for the academic year. If in school, please list your courses. If working, let us know something about the nature of your job.",
            },
            {
                "title": "Helpful Context",
                "limit": "No word limit",
                "themes": ["additional-info"],
                "text": "Please reflect on your primary application and share something not addressed elsewhere that would be helpful to the Admissions Committee as we review your file.",
            },
            {
                "title": "Why Geisel",
                "limit": "No word limit",
                "themes": ["why-school"],
                "text": "What aspects of the Geisel School of Medicine draw you to apply? Please include the characteristics and strengths you will bring to our program and how you hope to contribute to our community.",
            },
            {
                "title": "Being the Other",
                "limit": "250 words",
                "themes": ["diversity-equity"],
                "text": "Geisel School of Medicine values social justice and diversity in all its forms. Reflect on a situation where you were the 'other'.",
            },
        ],
    },
    {
        "name": "Tufts University School of Medicine",
        "slug": "tufts",
        "secondary_cycle": "2024 archive on Admit.org",
        "secondary_deadline": None,
        "official_url": "https://medicine.tufts.edu/admissions-financial-aid/admissions",
        "prompt_source": "https://med.admit.org/secondary-essays/tufts-university-school-of-medicine",
        "notes": "Tufts has a large prompt set but much of it maps cleanly to shared draft sources.",
        "prompts": [
            {
                "title": "Why Tufts",
                "limit": "1000 characters",
                "themes": ["why-school"],
                "text": "Do you wish to share a specific reason why have you chosen to apply to Tufts University School of Medicine?",
            },
            {
                "title": "Coming Year Plans",
                "limit": "1000 characters",
                "themes": ["gap-year"],
                "text": "Please briefly describe your plans for the coming year.",
            },
            {
                "title": "Journey to Medicine",
                "limit": "1000 characters",
                "themes": ["challenge", "diversity-equity"],
                "text": "Please tell us about your journey to medical school and how your background and experiences will positively impact your future as a medical student and physician.",
            },
            {
                "title": "Clinical Experience",
                "limit": "1000 characters",
                "themes": ["clinical"],
                "text": "Which of your experiences with clinical medicine or healthcare has best prepared you for a future career as a physician, and why?",
            },
            {
                "title": "Social Responsibility",
                "limit": "1000 characters",
                "themes": ["service"],
                "text": "Have you done substantial work or service that has contributed to societal good, or addressed social determinants of health, health equity, or social justice?",
            },
            {
                "title": "Withdrawals / Repeats",
                "limit": "1000 characters",
                "themes": ["academic-context"],
                "text": "Do you have any withdrawals or repeated coursework listed on your transcript(s)?",
            },
            {
                "title": "Leaves / Breaks",
                "limit": "1000 characters",
                "themes": ["academic-context"],
                "text": "Did you take any leaves of absence or significant breaks from your undergraduate education?",
            },
            {
                "title": "Academic Challenges",
                "limit": "1000 characters",
                "themes": ["academic-context"],
                "text": "We encourage you to use this space to elaborate on any academic challenges you have overcome since completing high school.",
            },
        ],
    },
    {
        "name": "The Ohio State University College of Medicine",
        "slug": "ohio-state",
        "secondary_cycle": "2024 archive on Admit.org search snippet",
        "secondary_deadline": "December 15, 2025",
        "official_url": "https://medicine.osu.edu/education/medical-school/admissions",
        "prompt_source": "https://med.admit.org/secondary-essays/ohio-state-university-college-of-medicine",
        "notes": "Track-specific prompts only matter if you apply to those tracks.",
        "prompts": [
            {
                "title": "Mission Areas",
                "limit": "250 words",
                "themes": ["mission-fit"],
                "text": "The mission statement of The Ohio State University College of Medicine is to improve people's lives through innovation in research, medical education, and patient care. Please describe how your past experiences predict your potential to contribute in two of these three areas.",
            },
            {
                "title": "Health Inequities",
                "limit": "250 words",
                "themes": ["diversity-equity"],
                "text": "Provide examples of factors contributing to health inequities that exist in the United States.",
            },
            {
                "title": "Primary Care Track Readiness",
                "limit": "250 words",
                "themes": ["future-goals"],
                "text": "Primary Care Track: How do you feel ready to choose a career in family medicine? Please provide any examples of exposure and/or experiences not already mentioned in your application.",
            },
            {
                "title": "Primary Care Track Rationale",
                "limit": "250 words",
                "themes": ["future-goals"],
                "text": "Primary Care Track: Describe your rationale for wanting to pursue an accelerated curriculum in preparation for a career in family medicine.",
            },
            {
                "title": "Community Medicine Track Rationale",
                "limit": "250 words",
                "themes": ["future-goals"],
                "text": "Community Medicine Track: Describe your rationale for wanting to pursue an accelerated curriculum in preparation for a career in family medicine.",
            },
            {
                "title": "Community Training Interest",
                "limit": "250 words",
                "themes": ["why-school", "service"],
                "text": "Community Medicine Track: Why are you interested in training in a smaller community, and what is the impact you hope to have by training in this setting?",
            },
        ],
    },
    {
        "name": "University of California San Diego School of Medicine",
        "slug": "ucsd",
        "secondary_cycle": "2026-2027 live secondary portal prompts provided by applicant on 2026-07-09",
        "secondary_deadline": None,
        "official_url": "https://medschool.ucsd.edu/education/md-program/admissions",
        "prompt_source": "https://applymed.ucsd.edu/security/login.aspx",
        "notes": "Updated from the applicant's UCSD secondary portal. Applicant pasted only the required autobiographical prompt, so program-specific prompts from the old archive were removed unless they are later selected in the portal.",
        "prompts": [
            {
                "title": "Autobiographical Statement",
                "limit": "6000 characters",
                "themes": ["challenge", "future-goals"],
                "text": "This should be a true autobiographical statement. Topics to be included are family, childhood, primary and secondary school years, undergraduate years, and, if applicable, what you’ve done since completing your bachelor’s degree. You should also discuss the motivational factors which led you to a career in medicine including any disadvantages or obstacles which might put your accomplishments into context. A repeat of your AMCAS statement will not be acceptable.",
            },
        ],
    },
    {
        "name": "Boston University Chobanian & Avedisian School of Medicine",
        "slug": "boston-university",
        "secondary_cycle": "2024 archive on Admit.org",
        "secondary_deadline": None,
        "official_url": "https://www.bumc.bu.edu/camed/admissions/",
        "prompt_source": "https://med.admit.org/secondary-essays/boston-university-school-of-medicine",
        "notes": "Several short explanatory prompts; good candidate for concise reusable context blocks.",
        "prompts": [
            {
                "title": "Direct to College",
                "limit": "2000 characters",
                "themes": ["academic-context"],
                "text": "Did you go on to college directly after high school?",
            },
            {
                "title": "Direct to Medical School / Gap Years",
                "limit": "2000 characters",
                "themes": ["gap-year"],
                "text": "Are you expecting to go on to medical school directly after completing your undergraduate degree? If you took gap year(s), please use this space to explain what you have been doing prior to applying to medical school.",
            },
            {
                "title": "More Than Four Years Undergraduate",
                "limit": "2000 characters",
                "themes": ["academic-context"],
                "text": "If you have spent more than 4 years as an undergraduate, please explain below.",
            },
            {
                "title": "Educational Narrative / Timeline",
                "limit": "2000 characters",
                "themes": ["additional-info"],
                "text": "Please provide a narrative or timeline to describe any features of your educational history that you think may be of particular interest to us.",
            },
            {
                "title": "Additional Strengths",
                "limit": "3000 characters",
                "themes": ["additional-info"],
                "text": "Use the space below to provide additional information you feel will provide us with a comprehensive understanding of your strengths as a candidate for a career in medicine.",
            },
            {
                "title": "Safety Net Environment",
                "limit": "3000 characters",
                "themes": ["why-school", "service"],
                "text": "Boston Medical Center is the largest safety net hospital in New England. Why are you specifically interested in beginning your medical education in this environment, and how do you feel that your previous experiences will prepare you for this unique learning environment?",
            },
        ],
    },
    {
        "name": "Kaiser Permanente Bernard J. Tyson School of Medicine",
        "slug": "kaiser-permanente",
        "secondary_cycle": "2026-2027 live secondary portal prompts provided by applicant on 2026-06-30",
        "secondary_deadline": "November 3, 2025",
        "official_url": "https://medschool.kp.org/admissions",
        "prompt_source": "https://applymedschool.kp.org/Security/Login.aspx",
        "notes": "Updated from the applicant's Kaiser Permanente secondary portal. Portal link added to secondary portal data.",
        "prompts": [
            {
                "title": "Unfavorable Outcome",
                "limit": "250 words",
                "themes": ["challenge"],
                "text": "During your career as a physician, you will likely encounter obstacles, and be required to overcome challenges. Please describe your experience with a situation that had an unfavorable outcome, including your reaction, how you might have responded differently, and what you learned about yourself.",
            },
            {
                "title": "Mission and Values Alignment",
                "limit": "250 words",
                "themes": ["mission-fit", "values"],
                "text": "Kaiser Permanente Bernard J. Tyson School of Medicine is dedicated to graduating courageous leaders who lead change through inquiry and innovation in medical education, the profession, and the healthcare system. How do your values align with this mission?",
            },
            {
                "title": "Lifelong Learning",
                "limit": "250 words",
                "themes": ["values"],
                "text": "Lifelong learning is an essential process for continued professional development. This includes reflection and being open and responsive to constructive feedback. Please tell us about an area of intellectual exploration you’re passionate about, and your approach to exploring this area.",
            },
            {
                "title": "Social Drivers and Community Advocacy",
                "limit": "250 words",
                "themes": ["service", "advocacy", "health-equity"],
                "text": "Drawing from your lived experiences, how do you think knowledge of the social drivers of health can be leveraged to engage in community health advocacy?",
            },
        ],
    },
    {
        "name": "Keck School of Medicine of USC",
        "slug": "usc-keck",
        "secondary_cycle": "2024 archive on Admit.org search snippet",
        "secondary_deadline": None,
        "official_url": "https://keck.usc.edu/education/doctor-of-medicine/admissions/",
        "prompt_source": "https://med.admit.org/secondary-essays/keck-school-of-medicine-of-the-university-of-southern-california",
        "notes": "Very short answers. Tone and precision matter more than exposition.",
        "prompts": [
            {
                "title": "Fun Lately",
                "limit": "65 words",
                "themes": ["additional-info"],
                "text": "What is the most fun you’ve had lately?",
            },
            {
                "title": "Nickname",
                "limit": "65 words",
                "themes": ["additional-info"],
                "text": "If you had to give yourself a nickname, what would it be?",
            },
            {
                "title": "Three Things You Don't Care About",
                "limit": "65 words",
                "themes": ["additional-info"],
                "text": "What are three things you don't care about at all?",
            },
            {
                "title": "Didn't Get What You Deserved",
                "limit": "65 words",
                "themes": ["challenge"],
                "text": "Describe a situation in which you didn’t get something you felt you deserved.",
            },
            {
                "title": "Anything Else",
                "limit": "200 words",
                "themes": ["additional-info"],
                "text": "Is there anything else you would like us to know?",
            },
            {
                "title": "Prior Medical School Application",
                "limit": "200 words",
                "themes": ["additional-info"],
                "text": "Have you previously applied to or attended medical school? If yes, please provide an explanation on what occurred with your application process or previous matriculation to medical school.",
            },
        ],
    },
    {
        "name": "University of Miami Leonard M. Miller School of Medicine",
        "slug": "miami-miller",
        "secondary_cycle": "2024 archive on Admit.org search snippet",
        "secondary_deadline": None,
        "official_url": "https://med.miami.edu/education/md-program/admissions",
        "prompt_source": "https://med.admit.org/secondary-essays/university-of-miami-leonard-m-miller-school-of-medicine",
        "notes": "Prompt text below reflects the search snippet surfaced from Admit.org on 2026-04-18.",
        "prompts": [
            {
                "title": "Why Miami",
                "limit": "5000 characters",
                "themes": ["why-school"],
                "text": "Why are you applying to the University of Miami Leonard M. Miller School of Medicine? Please provide a response that clearly articulates how you believe our program specifically will enhance your education/training, what you feel you will uniquely contribute to our learning community, and/or what features of our medical school prompted you to apply.",
            },
            {
                "title": "Diverse Environments",
                "limit": "5000 characters",
                "themes": ["diversity-equity"],
                "text": "Please discuss any experiences you may have had serving, working, living, and/or learning in broadly diverse environments that you believe would enable you to thrive in and contribute to our Miller SOM and Greater Miami community.",
            },
            {
                "title": "Values",
                "limit": "5000 characters",
                "themes": ["values"],
                "text": "Please discuss how your experiences align with core values such as integrity, collaboration, accountability, respect, and excellence.",
            },
        ],
    },
    {
        "name": "University of Virginia School of Medicine",
        "slug": "uva",
        "secondary_cycle": "2025-2026 from ProspectiveDoctor search snippet",
        "secondary_deadline": "December 15, 2025",
        "official_url": "https://med.virginia.edu/admissions/",
        "prompt_source": "https://www.prospectivedoctor.com/university-of-virginia-school-of-medicine-secondary/",
        "notes": "Using the latest public 2025-2026 prompt set surfaced in search results rather than the older Admit archive page.",
        "prompts": [
            {
                "title": "Why UVA",
                "limit": "350 words",
                "themes": ["why-school"],
                "text": "Why are you interested in attending the University of Virginia School of Medicine? What factors will be most important to you in choosing a medical school?",
            },
            {
                "title": "Challenge or Obstacle",
                "limit": "350 words",
                "themes": ["challenge"],
                "text": "Tell us about a challenge or obstacle you experienced. How did you manage it?",
            },
            {
                "title": "Pandemic Impact",
                "limit": "Short answer",
                "themes": ["additional-info"],
                "text": "Please briefly describe how the pandemic impacted your ability to pursue experiences like clinical work, shadowing, research, or community service.",
            },
            {
                "title": "Future Activities",
                "limit": "Short answer",
                "themes": ["gap-year"],
                "text": "Please list any and all future activities in clinical work, shadowing, research, or community service that you plan to pursue in the upcoming year.",
            },
        ],
    },
    {
        "name": "Wake Forest University School of Medicine",
        "slug": "wake-forest",
        "secondary_cycle": "2026-2027 live AMCAS portal PDF provided by applicant on 2026-06-26",
        "secondary_deadline": "November 30, 2025",
        "official_url": "https://school.wakehealth.edu/education-and-training/md-program/admissions",
        "prompt_source": "https://www.applyweb.com/cgi-bin/ustat?school=wfamca",
        "notes": "Updated from the applicant's Wake Forest AMCAS portal PDF. The portal warns not to disclose grades or MCAT scores and notes that reviewers do not have the full AMCAS personal statement and experiences while reviewing the secondary.",
        "prompts": [
            {
                "title": "Improve Health for All",
                "limit": "200 words",
                "themes": ["mission-fit", "service"],
                "text": "The mission of our academic health system is to improve health for all. Consider your abilities, insights, and previous experiences and then tell us about your potential to help improve health for all.",
            },
            {
                "title": "Non-Academic Challenge",
                "limit": "200 words",
                "themes": ["challenge"],
                "text": "Describe a non-academic challenge you have faced and explain how you overcame it.",
            },
            {
                "title": "Most Formative Experience",
                "limit": "200 words",
                "themes": ["clinical"],
                "text": "From your list of most meaningful experiences on the AMCAS application, choose one that has been the most formative in terms of your desire for a career in medicine. Why did that experience have such meaning for you in your decision-making process?",
            },
            {
                "title": "Collaboration",
                "limit": "200 words",
                "themes": ["teamwork"],
                "text": "Please share an experience that demonstrates how you have collaborated with others.",
            },
            {
                "title": "Commitment",
                "limit": "200 words",
                "themes": ["values"],
                "text": "Tell us about a time you have committed yourself to something.",
            },
            {
                "title": "Why Wake Forest",
                "limit": "150 words",
                "themes": ["why-school"],
                "text": "Tell us about any specific reason(s) (personal, educational, etc.) why you see yourself here at the Wake Forest University School of Medicine.",
            },
            {
                "title": "Interesting Fact",
                "limit": "50 words",
                "themes": ["additional-info"],
                "text": "Please tell us an interesting fact about yourself that a casual acquaintance may find surprising or interesting.",
            },
        ],
    },
    {
        "name": "University of California, Irvine School of Medicine",
        "slug": "uc-irvine",
        "secondary_cycle": "2024 archive on Admit.org search snippet",
        "secondary_deadline": None,
        "official_url": "https://www.meded.uci.edu/admissions/",
        "prompt_source": "https://med.admit.org/secondary-essays/university-of-california-irvine-school-of-medicine",
        "notes": "Search snippet exposed three prompts cleanly.",
        "prompts": [
            {
                "title": "Personal Accomplishment",
                "limit": "1500 characters",
                "themes": ["values"],
                "text": "What personal accomplishment are you most proud of and why?",
            },
            {
                "title": "Challenge or Disappointment",
                "limit": "1500 characters",
                "themes": ["challenge"],
                "text": "Please describe to the Admissions Committee a challenge or disappointment you have overcome and what you learned about yourself from that experience.",
            },
            {
                "title": "Marginalized Group",
                "limit": "1500 characters",
                "themes": ["diversity-equity"],
                "text": "Do you identify as being part of a marginalized group socioeconomically or in terms of access to quality education or healthcare? If so, please describe how this inequity has impacted you and your community.",
            },
        ],
    },
    {
        "name": "University of Maryland School of Medicine",
        "slug": "maryland",
        "secondary_cycle": "2025-2026 from Med School Insiders / ProspectiveDoctor search snippets",
        "secondary_deadline": "November 1, 2025 (MD); October 15, 2025 (MD/PhD)",
        "official_url": "https://www.medschool.umaryland.edu/admissions/",
        "prompt_source": "https://medschoolinsiders.com/medical-school-secondary-prompts-database/university-of-maryland-school-of-medicine-secondary-essay-prompts/",
        "notes": "Using a newer 2025-2026 set because the Admit.org page did not render beyond the JS-disabled notice.",
        "prompts": [
            {
                "title": "Relevant Background Characteristic",
                "limit": "200 characters",
                "themes": ["diversity-equity"],
                "text": "Please share any relevant characteristics about your background and/or experiences not already described in AMCAS that you feel may positively impact how you practice medicine in the future.",
            },
            {
                "title": "Why Maryland",
                "limit": "1000 characters",
                "themes": ["why-school"],
                "text": "Describe here the reason why you are specifically interested in attending the University of Maryland School of Medicine.",
            },
            {
                "title": "Current Academic Year",
                "limit": "1000 characters",
                "themes": ["gap-year"],
                "text": "Please describe what you will be doing during the academic year until you matriculate to medical school.",
            },
            {
                "title": "Academic Challenges",
                "limit": "1000 characters",
                "themes": ["academic-context"],
                "text": "If you’ve experienced academic challenges while in college and/or graduate or professional school, please describe and explain below.",
            },
            {
                "title": "Meaningful Clinical Exposure",
                "limit": "1500 characters",
                "themes": ["clinical"],
                "text": "Briefly describe your most meaningful exposure to clinical medicine.",
            },
            {
                "title": "Most Satisfying Community Service",
                "limit": "1500 characters",
                "themes": ["service"],
                "text": "Briefly describe your most satisfying experience related to community service.",
            },
            {
                "title": "Mission / Vision / Values Alignment",
                "limit": "1500 characters",
                "themes": ["mission-fit"],
                "text": "How does our School of Medicine mission, vision, and values align with your vision for your future career as a physician?",
            },
            {
                "title": "Entering a Profession",
                "limit": "1500 characters",
                "themes": ["values"],
                "text": "What does it mean to you to enter into a profession?",
            },
        ],
    },
    {
        "name": "Brown University Warren Alpert Medical School",
        "slug": "brown",
        "secondary_cycle": "2024 archive on Admit.org",
        "secondary_deadline": None,
        "official_url": "https://med.brown.edu/education/md-program/admissions",
        "prompt_source": "https://med.admit.org/secondary-essays/the-warren-alpert-medical-school-of-brown-university",
        "notes": "Compact prompt set; all three are strong candidates for reusable synced drafting.",
        "prompts": [
            {
                "title": "Current Year Activities",
                "limit": "2000 characters",
                "themes": ["gap-year"],
                "text": "Summarize your activities during the academic year. Describe how your activities are preparing you for a medical career.",
            },
            {
                "title": "Unique Attributes",
                "limit": "2000 characters",
                "themes": ["diversity-equity"],
                "text": "How will your unique attributes, life experiences, and interests add to the Alpert Medical School community?",
            },
            {
                "title": "Change Course",
                "limit": "3000 characters",
                "themes": ["challenge"],
                "text": "Reflect on a non-academic situation when you had to change course, and how you did so.",
            },
        ],
    },
    {
        "name": "Case Western Reserve University School of Medicine",
        "slug": "case-western",
        "secondary_cycle": "2024 archive on Admit.org",
        "secondary_deadline": None,
        "official_url": "https://case.edu/medicine/admissions-programs/md-programs/admissions",
        "prompt_source": "https://med.admit.org/secondary-essays/case-western-reserve-university-school-of-medicine",
        "notes": "This set covers challenge, regret, gap year, cultural humility, research, and open-ended context.",
        "prompts": [
            {
                "title": "Significant Personal Challenge",
                "limit": "2000 characters",
                "themes": ["challenge"],
                "text": "Please describe a significant personal challenge you have faced, one which you feel has helped to shape you as a person.",
            },
            {
                "title": "Would Handle Differently",
                "limit": "1000 characters",
                "themes": ["challenge"],
                "text": "Based on your current maturity and wisdom, reflect upon an experience from the past few years that you would handle differently today.",
            },
            {
                "title": "Gap Time",
                "limit": "1000 characters",
                "themes": ["gap-year"],
                "text": "If you are taking time off between college graduation and medical school matriculation, please tell us why you made this decision and what you will be doing or have done during this gap time.",
            },
            {
                "title": "Cultural Awareness and Humility",
                "limit": "1500 characters",
                "themes": ["diversity-equity"],
                "text": "Describe a personal experience that has deepened your cultural awareness and humility, and the impact of that learning on how you look at health and healthcare.",
            },
            {
                "title": "Research and Scholarship",
                "limit": "2500 characters",
                "themes": ["research"],
                "text": "Reflect on any research or other scholarly project, including how you learned from it, what challenged you, and how it may have impacted you educationally and professionally.",
            },
            {
                "title": "Further Information",
                "limit": "2000 characters",
                "themes": ["additional-info"],
                "text": "Is there any further information that you wish to share with the Admissions Committee that may not be captured in the rest of your application?",
            },
        ],
    },
    {
        "name": "University of Massachusetts Chan Medical School",
        "slug": "umass-chan",
        "secondary_cycle": "2026-2027 live secondary portal prompts provided by applicant on 2026-07-09",
        "secondary_deadline": None,
        "official_url": "https://www.umassmed.edu/som/admissions/",
        "prompt_source": "https://wsa-prd.erp.umasscs.net/psp/wsaprd92/EMPLOYEE/SA/?cmd=logout",
        "notes": "Updated from applicant-provided UMass Chan secondary prompts. Seven prompts captured; Q7 is optional and Q6 is reapplicant/N/A.",
        "prompts": [
            {
                "title": "Diverse Community Contribution",
                "limit": "200 words",
                "themes": ["diversity-equity"],
                "text": "How will your lived experiences, your background, your identity or that which makes you unique contribute to the diverse community and learning environment at UMass Chan?",
            },
            {
                "title": "Perseverance Through Challenging Circumstances",
                "limit": "200 words",
                "themes": ["challenge"],
                "text": "Describe a specific situation in which you persevered through more than usually challenging circumstances. What did you learn from that experience that you have continued to use?",
            },
            {
                "title": "Project, Solution, and Effectiveness",
                "limit": "200 words",
                "themes": ["research", "leadership"],
                "text": "Describe your role in a project that identified a problem, studied it, launched a solution and tested its effectiveness.",
            },
            {
                "title": "Self-Insight and Changed Approach",
                "limit": "200 words",
                "themes": ["values", "challenge"],
                "text": "Describe a situation where you learned something important about yourself that you did not previously know. How did you use that insight to change your approach in the future?",
            },
            {
                "title": "Systemic Inequity in Health Care",
                "limit": "200 words",
                "themes": ["diversity-equity", "service"],
                "text": "Please describe an example of your personal and/or professional experience with and understanding of systemic inequity, exclusion, or lack of representation in health care in the United States. How did you arrive at this understanding?",
            },
            {
                "title": "Reapplicant Candidacy Strengthening",
                "limit": "200 words",
                "themes": ["additional-info", "gap-year"],
                "text": "If you have previously applied to medical school, describe how you have strengthened your candidacy. If you have not previously applied to medical school please enter \"N/A\".",
            },
            {
                "title": "Optional Further Explanation",
                "limit": "200 words",
                "themes": ["academic-context", "additional-info"],
                "text": "Optional: Please discuss any part of your application that you believe requires further explanation such as low grades or scores or gaps in time that are not explained elsewhere.",
            },
        ],
    },
    {
        "name": "Albert Einstein College of Medicine",
        "slug": "albert-einstein",
        "secondary_cycle": "2024 archive on Admit.org search snippet",
        "secondary_deadline": None,
        "official_url": "https://einsteinmed.edu/education/md-program/admissions/",
        "prompt_source": "https://med.admit.org/secondary-essays/albert-einstein-college-of-medicine",
        "notes": "Admit search snippet exposed three prompts cleanly.",
        "prompts": [
            {
                "title": "Time Off During Undergraduate Years",
                "limit": "3000 characters",
                "themes": ["academic-context"],
                "text": "I have taken time off from school during my undergraduate years. If you answered yes to the above question, please explain.",
            },
            {
                "title": "Anything Else",
                "limit": "3000 characters",
                "themes": ["additional-info"],
                "text": "Please use this space to tell us anything about yourself that you would like us to know. If you do not wish to write anything, please write NA.",
            },
            {
                "title": "Unique Life Experiences",
                "limit": "3000 characters",
                "themes": ["diversity-equity", "challenge"],
                "text": "What unique life experiences, personal attributes and/or perspectives will you bring as part of the incoming class? Are there particular challenges or successes that you have encountered?",
            },
        ],
    },
    {
        "name": "VCU School of Medicine",
        "slug": "vcu",
        "secondary_cycle": "2024 archive on Admit.org",
        "secondary_deadline": None,
        "official_url": "https://medschool.vcu.edu/admissions/",
        "prompt_source": "https://med.admit.org/secondary-essays/virginia-commonwealth-university-school-of-medicine",
        "notes": "Very modular set; easy to draft in synced blocks.",
        "prompts": [
            {
                "title": "Mission and Values Alignment",
                "limit": "2000 characters",
                "themes": ["mission-fit"],
                "text": "Using specific examples, discuss how your mission and values align with those of VCU School of Medicine.",
            },
            {
                "title": "Grit",
                "limit": "2000 characters",
                "themes": ["challenge"],
                "text": "How do you define grit, and how have you demonstrated this in the past?",
            },
            {
                "title": "Lapses in Journey",
                "limit": "2000 characters",
                "themes": ["academic-context"],
                "text": "Please briefly explain any lapses in your journey to medicine that are not explained in your application.",
            },
            {
                "title": "Low GPA or Poor Grades",
                "limit": "2000 characters",
                "themes": ["academic-context"],
                "text": "Please briefly explain any low GPAs or poor grades.",
            },
            {
                "title": "What Are You Doing Now",
                "limit": "2000 characters",
                "themes": ["gap-year"],
                "text": "If not addressed in your application, what are you currently doing now?",
            },
        ],
    },
    {
        "name": "Indiana University School of Medicine",
        "slug": "indiana",
        "secondary_cycle": "2025-2026 non-essay secondary reporting from multiple advising sites",
        "secondary_deadline": "N/A",
        "official_url": "https://admissions.medicine.iu.edu/",
        "prompt_source": "https://www.prospectivedoctor.com/indiana-university-school-of-medicine-secondary/",
        "notes": "Multiple advising sources report that Indiana does not currently use essay-style secondary prompts; invited applicants complete additional non-essay information later in the process.",
        "prompts": [],
    },
]

_existing_school_map = {school["name"]: school for school in SCHOOLS}
_new_school_map = {school["name"]: school for school in NEW_SCHOOLS}
_merged_schools = []
_missing_school_names = []
for _school_name in ACTIVE_SCHOOL_NAMES:
    _school = _existing_school_map.get(_school_name) or _new_school_map.get(_school_name)
    if _school is None:
        _missing_school_names.append(_school_name)
    else:
        _merged_schools.append(_school)

if _missing_school_names:
    raise RuntimeError(f"Missing school metadata for: {', '.join(_missing_school_names)}")

SCHOOLS = _merged_schools


def slugify(text: str) -> str:
    value = re.sub(r"[^a-z0-9]+", "-", text.lower()).strip("-")
    return re.sub(r"-{2,}", "-", value)


def load_preliminary_stats() -> dict[str, dict[str, str]]:
    wb = load_workbook(SCHOOL_LIST_XLSX, data_only=True)
    ws = wb["MD Schools"]
    rows = list(ws.iter_rows(values_only=True))
    header_row = rows[3]
    header = [str(cell).strip() if cell is not None else "" for cell in header_row]
    school_idx = header.index("School")
    loc_idx = header.index("City")
    coa_idx = header.index("Estimated Annual COA ($)")
    gpa_idx = header.index("Reported GPA")
    mcat_idx = header.index("Reported MCAT")
    stats: dict[str, dict[str, str]] = {}
    for row in rows[4:]:
        if not row or row[school_idx] is None:
            continue
        name = str(row[school_idx]).strip()
        stats[name] = {
            "location": str(row[loc_idx]).strip() if row[loc_idx] is not None else "",
            "coa_per_year": str(row[coa_idx]).strip() if row[coa_idx] is not None else "",
            "median_gpa": str(row[gpa_idx]).strip() if row[gpa_idx] is not None else "",
            "median_mcat": str(row[mcat_idx]).strip() if row[mcat_idx] is not None else "",
        }
    return stats


def ensure_dir(path: Path) -> None:
    path.mkdir(parents=True, exist_ok=True)


def write_text(path: Path, text: str, *, overwrite: bool = True) -> None:
    if not overwrite and path.exists():
        return
    ensure_dir(path.parent)
    path.write_text(text.rstrip() + "\n", encoding="utf-8")


def build_primary_docs() -> None:
    primary_dir = ROOT / "essays" / "primary"
    write_text(
        primary_dir / "amcas-personal-statement.md",
        f"""# AMCAS Personal Statement

Source: 2027 AMCAS Applicant Guide (`{AMCAS_GUIDE}`)

## Hard Constraint
- {AMCAS_NOTES["personal_statement"]}

## Suggested Draft Structure
1. Opening scene or moment that clarifies why medicine became real for you.
2. Development through key clinical, service, research, or personal experiences.
3. Reflection on what changed in your understanding of physicianship.
4. Closing paragraph that points toward the physician you are trying to become.

## Draft Reminders
- The guide says the final submission should accurately reflect your own work and experiences.
- Do not repeat your secondary essays or activity descriptions word-for-word.
- Keep a "why medicine" through-line visible on every pass.
""",
    )
    write_text(
        primary_dir / "work-activities.md",
        f"""# AMCAS Work and Activities

Source: 2027 AMCAS Applicant Guide (`{AMCAS_GUIDE}`)

## Hard Constraints
- {AMCAS_NOTES["activities"]}
- AMCAS sends these descriptions to schools as plain text.

## Planning Grid
- Clinical
- Nonclinical service
- Research
- Leadership
- Teaching / mentoring
- Employment
- Awards / publications

## Draft Rule
Each entry should show:
1. What you did.
2. What mattered.
3. What changed in you.
""",
    )
    write_text(
        primary_dir / "most-meaningful-experiences.md",
        """# Most Meaningful Experiences

## Use This File To Decide
- Which experiences best explain why medicine
- Which experiences show growth, not just activity volume
- Which experiences you may want to reference again in secondaries

## For Each Experience
- Why this mattered more than your other entries
- What you learned about people, health, systems, or yourself
- How it still influences your path now
""",
    )
    write_text(
        primary_dir / "other-impactful-experiences.md",
        """# Other Impactful Experiences

Use this for context that changes how a reader should interpret your path.

## Good Uses
- Family or financial strain
- Caregiving
- Educational disruption
- Major personal context affecting opportunity or performance

## Keep The Focus On
- What the context was
- What it changed
- How you moved through it
""",
    )
    write_text(
        primary_dir / "amcas-personal-statement.draft.md",
        """# AMCAS Personal Statement Draft

Write your current working draft here.
""",
        overwrite=False,
    )
    write_text(
        primary_dir / "most-meaningful-experiences.draft.md",
        """# Most Meaningful Experiences Draft

Use this file for the narrative versions of your three most meaningful experiences.
""",
        overwrite=False,
    )
    write_text(
        primary_dir / "other-impactful-experiences.draft.md",
        """# Other Impactful Experiences Draft

Write any contextual impact statement drafts here.
""",
        overwrite=False,
    )


def build_process_docs() -> None:
    write_text(
        ROOT / "docs" / "process" / "amcas-2027-guide.md",
        f"""# AMCAS 2027 Guide Notes

Primary source: `{AMCAS_GUIDE}`

## Verified Notes From The Guide
- {AMCAS_NOTES["personal_statement"]}
- {AMCAS_NOTES["verification"]}
- {AMCAS_NOTES["transcripts"]}
- {AMCAS_NOTES["letters"]}
- {AMCAS_NOTES["activities"]}

## Operational Takeaways
- Submit AMCAS as early as your primary is truly ready; verification timing is the main bottleneck.
- Track transcript arrival separately from letter arrival.
- Draft secondaries before invitations land, especially for your highest-priority schools.
- Preserve reusable language in shared draft sources, then customize final school drafts directly.
""",
    )
    write_text(
        ROOT / "docs" / "process" / "application-workflow.md",
        """# Application Workflow

## Recommended Order
1. Finish AMCAS personal statement and Work/Activities.
2. Write actual drafts in `*.draft.md` files so regeneration never overwrites your essay text.
3. Edit shared draft sources in `essays/shared-drafts/` when multiple schools use the same answer.
4. Run `python3 scripts/sync_shared_essay_drafts.py` after school-list or prompt changes.
5. Render clean review pages from your markdown drafts with `python3 scripts/render_review_site.py`.

## Editing Rules For This Repo
- Write your actual essay answer in `prompt-XX.draft.md`.
- If a school draft is a symlink, editing it also edits the shared source in `essays/shared-drafts/`.
- Prompt titles and limits come from `data/schools.json`.
""",
    )
    write_text(
        ROOT / "docs" / "research" / "sources.md",
        "\n".join(
            [
                "# Sources",
                "",
                f"Repository build date: {BUILD_DATE}",
                "",
                "These are the prompt/archive pages and admissions pages used to seed this workspace.",
                "",
            ]
            + [f"- {school['name']}: {school['prompt_source']}" for school in SCHOOLS]
        ),
    )


def build_school_files(stats: dict[str, dict[str, str]]) -> list[dict[str, object]]:
    exported = []
    for school in SCHOOLS:
        school_stats = stats.get(school["name"], {})
        school_dir = ROOT / "schools" / school["slug"]
        ensure_dir(school_dir / "essays")

        prompt_list = school["prompts"]
        school_readme_lines = [
            f"# {school['name']}",
            "",
            f"- Active list location: {school_stats.get('location', 'N/A')}",
            f"- Active list COA/year: {school_stats.get('coa_per_year', 'N/A')}",
            f"- Active list GPA median: {school_stats.get('median_gpa', 'N/A')}",
            f"- Active list MCAT median: {school_stats.get('median_mcat', 'N/A')}",
            f"- Latest prompt cycle used here: {school['secondary_cycle']}",
            f"- Latest verified secondary deadline: {school['secondary_deadline'] or 'See admissions source'}",
            f"- Official admissions page: {school['official_url']}",
            f"- Prompt source: {school['prompt_source']}",
            "",
            "## Required Core Items",
        ]
        school_readme_lines.extend([f"- {item}" for item in PRIMARY_REQUIREMENTS])
        school_readme_lines.extend(
            [
                "",
                "## School-Specific Notes",
                f"- {school['notes']}",
                f"- Secondary prompt count captured in this repo: {len(prompt_list)}",
                "",
                "## Essay Drafts",
            ]
        )
        if not prompt_list:
            school_readme_lines.append("- No essay-style secondary prompts captured; keep tracking this school's non-essay secondary requirements.")

        exported_school = {
            "name": school["name"],
            "slug": school["slug"],
            "location": school_stats.get("location", ""),
            "coa_per_year": school_stats.get("coa_per_year", ""),
            "median_gpa": school_stats.get("median_gpa", ""),
            "median_mcat": school_stats.get("median_mcat", ""),
            "secondary_cycle": school["secondary_cycle"],
            "secondary_deadline": school["secondary_deadline"],
            "official_url": school["official_url"],
            "prompt_source": school["prompt_source"],
            "notes": school["notes"],
            "prompts": [],
        }

        for idx, prompt in enumerate(prompt_list, start=1):
            draft_name = f"prompt-{idx:02d}.draft.md"
            draft_path = school_dir / "essays" / draft_name
            write_text(
                draft_path,
                f"""# {school['name']} - {prompt['title']} Draft

Write your current draft here.
""",
                overwrite=False,
            )

            school_readme_lines.append(f"- `{draft_path.relative_to(ROOT)}`")
            exported_school["prompts"].append(prompt)

        write_text(school_dir / "README.md", "\n".join(school_readme_lines))
        exported.append(exported_school)
    return exported


def build_root_readme() -> None:
    school_count = len(SCHOOLS)
    write_text(
        ROOT / "README.md",
        f"""# Med School Applications Workspace

This repository was generated on {BUILD_DATE} to support your VS Code + Codex workflow for the 2027 AMCAS cycle.

## What This Repo Does
- Normalizes your active school list into a reusable local dataset
- Creates primary essay planning files
- Creates school-by-school secondary draft placeholders
- Links functionally similar essay drafts through `essays/shared-drafts/`
- Renders clean HTML review pages from Markdown drafts

## Quick Start
1. Draft primary materials in `essays/primary/`.
2. Write actual essay drafts in `*.draft.md`. Some are linked to shared source drafts in `essays/shared-drafts/`.
3. Refresh shared essay links after school-list or prompt changes with:

```bash
python3 scripts/sync_shared_essay_drafts.py
```
4. Render polished review pages from your markdown drafts with:

```bash
python3 scripts/render_review_site.py
```

## Important Notes
- Prompt sets in this repo use the latest public sources I could verify on {BUILD_DATE}. Some schools are using 2024 Admit.org archives; a smaller number use newer 2025-2026 prompts from alternate public advising sources where Admit.org did not expose the full prompt text.
- Treat prompt metadata as a drafting aid, not as a substitute for checking each school’s live portal when secondaries open.
- The active MD school list currently includes {school_count} schools.
""",
    )


def main() -> None:
    ensure_dir(ROOT / "scripts")
    build_primary_docs()
    build_process_docs()
    stats = load_preliminary_stats()
    exported = build_school_files(stats)
    write_text(ROOT / "data" / "schools.json", json.dumps(exported, indent=2))
    build_root_readme()


if __name__ == "__main__":
    main()

import json
import re
from copy import copy
from pathlib import Path

from openpyxl import load_workbook


WORKBOOK_PATH = "med_school_list_md.xlsx"
ACTIVE_SCHOOLS_PATH = Path("data/schools.json")


SELECTED_SCHOOLS = [
    {
        "school": "UC Davis School of Medicine",
        "bucket": "Safety",
        "mission": "Transform lives by improving health through education, research, clinical care, and community.",
        "notes": "Strong in-state value with clear community-health and service alignment for your hospice, clinic, and justice-oriented work.",
    },
    {
        "school": "University of California San Diego School of Medicine",
        "bucket": "Target",
        "mission": "Serve communities and create a healthier world through research, medical education, and clinical care.",
        "notes": "Good California fit for a research-active applicant who also brings teaching and patient-facing experience.",
    },
    {
        "school": "Kaiser Permanente Bernard J. Tyson School of Medicine",
        "bucket": "Target",
        "mission": "Provide a world-class medical education that ignites service and improves the health of patients and communities.",
        "notes": "Excellent fit for systems thinking, health equity, student leadership, and service-centered clinical motivation.",
    },
    {
        "school": "Keck School of Medicine of USC",
        "bucket": "Reach",
        "mission": "Ignite discovery through education, research, and patient care to optimize health for diverse communities.",
        "notes": "Los Angeles academic-medicine option with strong research resources and diverse-community care exposure.",
    },
    {
        "school": "University of Rochester School of Medicine and Dentistry",
        "bucket": "Target",
        "mission": "Improve the health of individuals and populations through innovative medical education grounded in humility and personal growth.",
        "notes": "Biopsychosocial and reflective training model fits hospice work, mentoring, and thoughtful service.",
    },
    {
        "school": "University of Colorado School of Medicine",
        "bucket": "Target",
        "mission": "Deliver excellence in education, research, clinical care, and community collaboration to advance health equity.",
        "notes": "Balanced science-plus-service mission fits your microbiology research, teaching, and community engagement.",
    },
    {
        "school": "Lewis Katz School of Medicine at Temple University",
        "bucket": "Safety",
        "mission": "Align with diverse communities to advance medicine and improve health through education, research, and training.",
        "notes": "One of the cleanest mission fits for your urban-service, hospice, and restorative justice experiences.",
    },
    {
        "school": "Emory University School of Medicine",
        "bucket": "Target",
        "mission": "Pursue the highest standards in education, biomedical research, and patient care while preparing leaders to serve diverse communities.",
        "notes": "Research-forward but still service-conscious; your poster output and community work both help here.",
    },
    {
        "school": "University of Cincinnati College of Medicine",
        "bucket": "Safety",
        "mission": "Educate and train future physicians, advance knowledge through innovative research, and improve health through compassionate patient-centered care.",
        "notes": "Nice balance of attainable stats, collaborative culture, and room for your research plus clinical support background.",
    },
    {
        "school": "Geisel School of Medicine at Dartmouth",
        "bucket": "Target",
        "mission": "Address the world's health problems through research, health-system improvement, and education of future physicians and scientists.",
        "notes": "Good fit for research curiosity paired with systems-level service and equity interests.",
    },
    {
        "school": "Tufts University School of Medicine",
        "bucket": "Target",
        "mission": "Improve the health and well-being of communities through education, advancing knowledge, and compassionate patient care.",
        "notes": "Humanistic and community-oriented mission fits your teaching experience and long-term service commitments.",
    },
    {
        "school": "The Ohio State University College of Medicine",
        "bucket": "Target",
        "mission": "Transform the health of communities through inclusive and innovative education, discovery, and care.",
        "notes": "Inclusive academic-medicine environment that matches your leadership, research, and patient-centered service profile.",
    },
    {
        "school": "Boston University Chobanian & Avedisian School of Medicine",
        "bucket": "Reach",
        "mission": "Cultivate learners from varied backgrounds committed to innovation, medicine, research, and the well-being of all communities.",
        "notes": "Urban mission and social-justice emphasis fit well, but the stats keep this in the reach tier.",
    },
    {
        "school": "Virginia Tech Carilion School of Medicine",
        "bucket": "Safety",
        "mission": "Prepare physician thought leaders through innovations in medical education and cutting-edge discovery to improve community health.",
        "notes": "Great match for your research output, poster activity, and leadership narrative.",
    },
    {
        "school": "University of Virginia School of Medicine",
        "bucket": "Reach",
        "mission": "Transform health and inspire hope through research, education, patient care, and community partnership.",
        "notes": "Still a reach, but your service record and conference-backed research make the fit coherent.",
    },
    {
        "school": "Vermont Larner College of Medicine",
        "degree": "MD",
        "city": "Burlington, VT",
        "coa": 100036,
        "gpa": 3.75,
        "mcat": 512,
        "bucket": "Target",
        "coa_basis": "Public (out-of-state COA)",
        "stats_url": "https://www.medcmp.com/university-of-vermont-college-of-medicine/acceptance-rate-and-mcat-scores/",
        "coa_url": "https://www.medcmp.com/university-of-vermont-college-of-medicine/cost-of-attendance/",
        "mission": "Educate skilled physicians and scientists who improve health through patient-centered care, scholarship, and service.",
        "notes": "Supportive, humanistic training environment with strong service and professionalism overlap for your profile.",
    },
    {
        "school": "University of Wisconsin School of Medicine and Public Health",
        "degree": "MD",
        "city": "Madison, WI",
        "coa": 94424,
        "gpa": 3.78,
        "mcat": 511,
        "bucket": "Target",
        "coa_basis": "Public (out-of-state COA)",
        "stats_url": "https://www.medcmp.com/university-of-wisconsin-school-of-medicine-and-public-health/acceptance-rate-and-mcat-scores/",
        "coa_url": "https://www.medcmp.com/university-of-wisconsin-school-of-medicine-and-public-health/cost-of-attendance/",
        "mission": "Advance health without compromise through service, scholarship, science, and social responsibility.",
        "notes": "Public-health and social-responsibility framing fits your community work better than a pure lab-heavy school would.",
    },
    {
        "school": "Medical College of Wisconsin",
        "degree": "MD",
        "city": "Milwaukee, WI",
        "coa": 101344,
        "gpa": 3.80,
        "mcat": 510,
        "bucket": "Target",
        "coa_basis": "Private",
        "stats_url": "https://www.medcmp.com/medical-college-of-wisconsin/acceptance-rate-and-mcat-scores/",
        "coa_url": "https://www.medcmp.com/medical-college-of-wisconsin/cost-of-attendance/",
        "mission": "Build a healthier world through discovery, patient care, community engagement, and the education of tomorrow's healers.",
        "notes": "Strong match for research plus community engagement, without demanding ultra-reach stats.",
    },
    {
        "school": "Sidney Kimmel Medical College at Thomas Jefferson University",
        "degree": "MD",
        "city": "Philadelphia, PA",
        "coa": 90759,
        "gpa": 3.80,
        "mcat": 513,
        "bucket": "Target",
        "coa_basis": "Private",
        "stats_url": "https://www.medcmp.com/sidney-kimmel-medical-college-at-thomas-jefferson-university/acceptance-rate-and-mcat-scores/",
        "coa_url": "https://www.medcmp.com/sidney-kimmel-medical-college-at-thomas-jefferson-university/cost-of-attendance/",
        "mission": "Improve lives through compassionate care, innovation, education, and service to diverse communities.",
        "notes": "Urban patient population, inclusive values, and strong clinical training make this a sensible fit school.",
    },
    {
        "school": "University of Iowa Roy J. and Lucille A. Carver College of Medicine",
        "degree": "MD",
        "city": "Iowa City, IA",
        "coa": 92614,
        "gpa": 3.81,
        "mcat": 514,
        "bucket": "Target",
        "coa_basis": "Public (out-of-state COA)",
        "stats_url": "https://www.medcmp.com/university-of-iowa-roy-j-and-lucille-a-carver-college-of-medicine/acceptance-rate-and-mcat-scores/",
        "coa_url": "https://www.medcmp.com/university-of-iowa-roy-j-and-lucille-a-carver-college-of-medicine/cost-of-attendance/",
        "mission": "Inspire and educate world-class health care providers and scientists for Iowa and the global community.",
        "notes": "Case-based, research-capable program with enough service orientation to fit your broader application story.",
    },
    {
        "school": "Western Michigan University Homer Stryker M.D. School of Medicine",
        "degree": "MD",
        "city": "Kalamazoo, MI",
        "coa": 100715,
        "gpa": 3.81,
        "mcat": 512,
        "bucket": "Target",
        "coa_basis": "Private",
        "stats_url": "https://www.medcmp.com/western-michigan-university-homer-stryker-md-school-of-medicine/acceptance-rate-and-mcat-scores/",
        "coa_url": "https://www.medcmp.com/western-michigan-university-homer-stryker-md-school-of-medicine/cost-of-attendance/",
        "mission": "Develop exceptional and compassionate physicians through patient-centered, discovery-driven medical education.",
        "notes": "Patient-centered training plus research space makes this a balanced fit for your profile.",
    },
    {
        "school": "University of Miami Leonard M. Miller School of Medicine",
        "degree": "MD",
        "city": "Miami, FL",
        "coa": 106785,
        "gpa": 3.85,
        "mcat": 516,
        "bucket": "Reach",
        "coa_basis": "Private",
        "stats_url": "https://www.medcmp.com/university-of-miami-leonard-m-miller-school-of-medicine/acceptance-rate-and-mcat-scores/",
        "coa_url": "https://www.medcmp.com/university-of-miami-leonard-m-miller-school-of-medicine/cost-of-attendance/",
        "mission": "Advance medical education, discovery, compassionate care, and community health across diverse populations.",
        "notes": "Diverse-community focus helps, but the academic bar is still high enough to keep it as a reach.",
    },
    {
        "school": "Wake Forest University School of Medicine",
        "degree": "MD",
        "city": "Winston-Salem, NC",
        "coa": 99067,
        "gpa": 3.85,
        "mcat": 512,
        "bucket": "Target",
        "coa_basis": "Private",
        "stats_url": "https://www.medcmp.com/wake-forest-school-of-medicine/acceptance-rate-and-mcat-scores/",
        "coa_url": "https://www.medcmp.com/wake-forest-school-of-medicine/cost-of-attendance/",
        "mission": "Generate and translate knowledge to improve health through research, learning, and compassionate care.",
        "notes": "Research and service both matter here, which helps your posters feel connected to patient-facing work.",
    },
    {
        "school": "Albany Medical College",
        "degree": "MD",
        "city": "Albany, NY",
        "coa": 85749,
        "gpa": 3.80,
        "mcat": 510,
        "bucket": "Safety",
        "coa_basis": "Private",
        "stats_url": "https://www.medcmp.com/albany-medical-college/acceptance-rate-and-mcat-scores/",
        "coa_url": "https://www.medcmp.com/albany-medical-college/cost-of-attendance/",
        "mission": "Improve health through biomedical science, patient-centered medicine, and service to communities.",
        "notes": "Smaller, humane, clinically grounded environment that still values science and service.",
    },
    {
        "school": "Drexel University College of Medicine",
        "degree": "MD",
        "city": "Philadelphia, PA",
        "coa": 106067,
        "gpa": 3.76,
        "mcat": 512,
        "bucket": "Safety",
        "coa_basis": "Private",
        "stats_url": "https://www.medcmp.com/drexel-university-college-of-medicine/acceptance-rate-and-mcat-scores/",
        "coa_url": "https://www.medcmp.com/drexel-university-college-of-medicine/cost-of-attendance/",
        "mission": "Educate a diverse physician workforce committed to innovation, humanism, and community health.",
        "notes": "The blend of opportunity, inclusion, and service makes this one of the better mission fits on the list.",
    },
    {
        "school": "Wayne State University School of Medicine",
        "degree": "MD",
        "city": "Detroit, MI",
        "coa": 101999,
        "gpa": 3.78,
        "mcat": 511,
        "bucket": "Safety",
        "coa_basis": "Public (out-of-state COA)",
        "stats_url": "https://www.medcmp.com/wayne-state-university-school-of-medicine/acceptance-rate-and-mcat-scores/",
        "coa_url": "https://www.medcmp.com/wayne-state-university-school-of-medicine/cost-of-attendance/",
        "mission": "Educate a diverse student body in an urban setting and culture of inclusion to prepare leaders who achieve health and wellness for society.",
        "notes": "Urban mission, inclusion focus, and community investment match your justice-oriented service especially well.",
    },
    {
        "school": "Loyola University Chicago Stritch School of Medicine",
        "degree": "MD",
        "city": "Maywood, IL",
        "coa": 70998,
        "gpa": 3.80,
        "mcat": 512,
        "bucket": "Target",
        "coa_basis": "Private (medcmp listed total matches tuition; verify full school budget)",
        "stats_url": "https://www.medcmp.com/loyola-university-chicago-stritch-school-of-medicine/acceptance-rate-and-mcat-scores/",
        "coa_url": "https://www.medcmp.com/loyola-university-chicago-stritch-school-of-medicine/cost-of-attendance/",
        "mission": "Educate physicians and scientists in a Jesuit tradition of scholarship, innovation, service, and respect for human dignity.",
        "notes": "Whole-person care, social justice, and service to others line up very naturally with your volunteering record.",
    },
    {
        "school": "Frank H. Netter M.D. School of Medicine at Quinnipiac University",
        "degree": "MD",
        "city": "North Haven, CT",
        "coa": 95997,
        "gpa": 3.78,
        "mcat": 513,
        "bucket": "Target",
        "coa_basis": "Private",
        "stats_url": "https://www.medcmp.com/frank-h-netter-md-school-of-medicine-at-quinnipiac-university/acceptance-rate-and-mcat-scores/",
        "coa_url": "https://www.medcmp.com/frank-h-netter-md-school-of-medicine-at-quinnipiac-university/cost-of-attendance/",
        "mission": "Prepare empathetic, inclusive physicians who improve health in collaboration with patients and communities.",
        "notes": "Patient-centered, collaborative culture makes sense for your teaching and direct-service strengths.",
    },
    {
        "school": "Tulane University School of Medicine",
        "degree": "MD",
        "city": "New Orleans, LA",
        "coa": 95328,
        "gpa": 3.70,
        "mcat": 510,
        "bucket": "Safety",
        "coa_basis": "Private",
        "stats_url": "https://www.medcmp.com/tulane-university-school-of-medicine/acceptance-rate-and-mcat-scores/",
        "coa_url": "https://www.medcmp.com/tulane-university-school-of-medicine/cost-of-attendance/",
        "mission": "Improve human health and foster healthy communities through discovery, education, and high-quality patient care.",
        "notes": "Public-facing mission, underserved-community work, and infectious-disease history fit your application themes well.",
    },
    {
        "school": "Eastern Virginia Medical School",
        "degree": "MD",
        "city": "Norfolk, VA",
        "coa": 86585,
        "gpa": 3.80,
        "mcat": 513,
        "bucket": "Target",
        "coa_basis": "Public (out-of-state COA)",
        "stats_url": "https://www.medcmp.com/eastern-virginia-medical-school/acceptance-rate-and-mcat-scores/",
        "coa_url": "https://www.medcmp.com/eastern-virginia-medical-school/cost-of-attendance/",
        "mission": "Improve community and global health through education, research, and patient-centered care.",
        "notes": "Community-oriented training and broad clinical service ethos fit your existing experience well.",
    },
    {
        "school": "Rush Medical College at Rush University",
        "degree": "MD",
        "city": "Chicago, IL",
        "coa": 104682,
        "gpa": 3.72,
        "mcat": 510,
        "bucket": "Safety",
        "coa_basis": "Private",
        "stats_url": "https://www.medcmp.com/rush-medical-college/acceptance-rate-and-mcat-scores/",
        "coa_url": "https://www.medcmp.com/rush-medical-college/cost-of-attendance/",
        "mission": "Train clinically excellent physicians committed to service, equity, and outstanding care for diverse urban communities.",
        "notes": "One of the strongest pure service-fit schools for your hospice and restorative justice background.",
    },
    {
        "school": "Chicago Medical School at Rosalind Franklin University of Medicine and Science",
        "degree": "MD",
        "city": "North Chicago, IL",
        "coa": 100423,
        "gpa": 3.69,
        "mcat": 508,
        "bucket": "Safety",
        "coa_basis": "Private",
        "stats_url": "https://www.medcmp.com/chicago-medical-school-of-rosalind-franklin-university-of-medicine-and-science/acceptance-rate-and-mcat-scores/",
        "coa_url": "https://www.medcmp.com/chicago-medical-school-of-rosalind-franklin-university-of-medicine-and-science/cost-of-attendance/",
        "mission": "Prepare patient-centered physicians and scientists dedicated to discovery, respect, and service.",
        "notes": "Stats fit is favorable, and the patient-centered mission still connects well to your experiences.",
    },
    {
        "school": "New York Medical College",
        "degree": "MD",
        "city": "Valhalla, NY",
        "coa": 94496,
        "gpa": 3.80,
        "mcat": 516,
        "bucket": "Target",
        "coa_basis": "Private",
        "stats_url": "https://www.medcmp.com/new-york-medical-college/acceptance-rate-and-mcat-scores/",
        "coa_url": "https://www.medcmp.com/new-york-medical-college/cost-of-attendance/",
        "mission": "Educate skilled, ethical physicians and scientists who advance health through scholarship and compassionate care.",
        "notes": "A solid humanistic academic-medicine option; your service keeps it mission-consistent despite the slightly higher MCAT.",
    },
    {
        "school": "University of California, Riverside School of Medicine",
        "degree": "MD",
        "city": "Riverside, CA",
        "coa": 66443,
        "gpa": 3.60,
        "mcat": 509,
        "bucket": "Safety",
        "coa_basis": "CA public (in-state COA)",
        "stats_url": "https://www.medcmp.com/university-of-california-riverside-school-of-medicine/acceptance-rate-and-mcat-scores/",
        "coa_url": "https://www.medcmp.com/university-of-california-riverside-school-of-medicine/cost-of-attendance/",
        "mission": "Improve the health of California, especially Inland Southern California, by training a diverse physician workforce and serving the medically underserved.",
        "notes": "Mission fit is excellent for service and California preference, even without a deep Inland Empire tie.",
    },
    {
        "school": "California University of Science and Medicine",
        "degree": "MD",
        "city": "Colton, CA",
        "coa": None,
        "gpa": 3.60,
        "mcat": 513,
        "bucket": "Target",
        "coa_basis": "Private (verify official COA)",
        "stats_url": "https://www.medcmp.com/california-university-of-science-and-medicine/acceptance-rate-and-mcat-scores/",
        "coa_url": "https://www.medcmp.com/california-university-of-science-and-medicine/cost-of-attendance/",
        "mission": "Develop excellent and caring physicians, scientists, and leaders who improve community health and expand opportunity, especially for California and the Inland Empire.",
        "notes": "I included this because it adds a sixth California MD option with a strong underserved-community mission and favorable GPA fit.",
    },
]


def clone_style(src, dst):
    dst._style = copy(src._style)
    if src.number_format:
        dst.number_format = src.number_format
    if src.alignment:
        align = copy(src.alignment)
        align.wrap_text = True
        dst.alignment = align


def normalize_key(value):
    return re.sub(r"[^a-z0-9]+", " ", str(value or "").lower()).strip()


def number_or_blank(value, cast):
    if value in (None, ""):
        return None
    try:
        return cast(value)
    except (TypeError, ValueError):
        return value


def load_active_school_rows():
    if not ACTIVE_SCHOOLS_PATH.exists():
        return {}
    rows = json.loads(ACTIVE_SCHOOLS_PATH.read_text())
    return {normalize_key(row["name"]): row for row in rows}


def build_row(entry, existing_rows, active_rows):
    active = active_rows.get(normalize_key(entry["school"]), {})
    row = dict(existing_rows.get(entry["school"]) or existing_rows.get(normalize_key(entry["school"])) or {})

    row["School"] = entry["school"]
    row["Degree"] = entry.get("degree") or row.get("Degree") or "MD"
    row["City"] = entry.get("city") or row.get("City") or active.get("location")
    row["Estimated Annual COA ($)"] = entry.get("coa", row.get("Estimated Annual COA ($)"))
    if row["Estimated Annual COA ($)"] in (None, ""):
        row["Estimated Annual COA ($)"] = number_or_blank(active.get("coa_per_year"), int)
    row["Reported GPA"] = entry.get("gpa", row.get("Reported GPA"))
    if row["Reported GPA"] in (None, ""):
        row["Reported GPA"] = number_or_blank(active.get("median_gpa"), float)
    row["Reported MCAT"] = entry.get("mcat", row.get("Reported MCAT"))
    if row["Reported MCAT"] in (None, ""):
        row["Reported MCAT"] = number_or_blank(active.get("median_mcat"), int)
    row["Reach/Target/Safety"] = entry["bucket"]
    row["COA Basis"] = entry.get("coa_basis") or row.get("COA Basis")
    row["Stats Source URL"] = entry.get("stats_url") or row.get("Stats Source URL") or active.get("official_url")
    row["COA Source URL"] = entry.get("coa_url") or row.get("COA Source URL")
    row["Notes"] = entry["notes"]
    row["Mission Statement"] = entry["mission"]
    return row


def main():
    wb = load_workbook(WORKBOOK_PATH)
    ws = wb["MD Schools"]
    summary = wb["Summary"]

    headers = [
        "School",
        "Degree",
        "City",
        "Estimated Annual COA ($)",
        "Reported GPA",
        "Reported MCAT",
        "Reach/Target/Safety",
        "COA Basis",
        "Stats Source URL",
        "COA Source URL",
        "Notes",
        "Mission Statement",
    ]

    if ws.max_column < 12:
        ws.insert_cols(12, 1)
        for row_idx in range(1, ws.max_row + 1):
            clone_style(ws.cell(row_idx, 11), ws.cell(row_idx, 12))

    existing_rows = {}
    old_headers = [ws.cell(4, col).value for col in range(1, 12)]
    for row_idx in range(5, ws.max_row + 1):
        school = ws.cell(row_idx, 1).value
        if not school:
            continue
        row_values = [ws.cell(row_idx, col).value for col in range(1, 12)]
        row = dict(zip(old_headers, row_values))
        existing_rows[school] = row
        existing_rows[normalize_key(school)] = row

    active_rows = load_active_school_rows()
    selected_rows = [build_row(entry, existing_rows, active_rows) for entry in SELECTED_SCHOOLS]

    for merged in list(ws.merged_cells.ranges):
        if str(merged) in {"A1:K1", "A2:K2", "A1:L1", "A2:L2"}:
            ws.unmerge_cells(str(merged))
    ws.merge_cells("A1:L1")
    ws.merge_cells("A2:L2")

    ws["A1"] = "U.S. MD school list tailored for 3.73 GPA / 520 MCAT / California resident / GPA cap 3.85"
    ws["A2"] = (
        "MD-only schools with reported GPA at or below 3.85. Mission fit was weighted toward research, "
        "hospice/service, restorative justice/community work, teaching, and student leadership, with extra weight on California options."
    )

    for col_idx, header in enumerate(headers, start=1):
        ws.cell(4, col_idx).value = header

    max_rows_to_clear = max(ws.max_row, 4 + len(selected_rows) + 20)
    for row_idx in range(5, max_rows_to_clear + 1):
        for col_idx in range(1, 13):
            ws.cell(row_idx, col_idx).value = None

    template_row = 5
    for offset, row in enumerate(selected_rows):
        row_idx = 5 + offset
        if row_idx != template_row and template_row <= ws.max_row:
            for col_idx in range(1, 13):
                clone_style(ws.cell(template_row, col_idx), ws.cell(row_idx, col_idx))
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row_idx, col_idx).value = row.get(header)

    for row_idx in range(4, 5 + len(selected_rows)):
        for col_idx in (11, 12):
            align = copy(ws.cell(row_idx, col_idx).alignment)
            align.wrap_text = True
            ws.cell(row_idx, col_idx).alignment = align

    ws.column_dimensions["A"].width = 44
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 18
    ws.column_dimensions["H"].width = 28
    ws.column_dimensions["I"].width = 42
    ws.column_dimensions["J"].width = 42
    ws.column_dimensions["K"].width = 62
    ws.column_dimensions["L"].width = 76
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:L{4 + len(selected_rows)}"

    summary["A1"] = "Selection notes"
    summary["A3"] = "This version stays at 35 schools, is MD-only, and excludes schools above a reported 3.85 GPA."
    summary["A4"] = "Mission fit was screened against your resume themes: research, hospice/service, teaching, leadership, and justice-oriented community work."
    summary["A5"] = "I leaned harder toward California schools where the stats and mission still made sense, including UC Riverside and CUSM."
    summary["A6"] = "For MD admissions, 'Safety' still means lower relative risk, not a guarantee."

    summary["A9"] = "Count by bucket"
    summary["D9"] = "Average reported stats by bucket"
    summary["A10"] = "Safety"
    summary["A11"] = "Target"
    summary["A12"] = "Reach"
    summary["B10"] = '=COUNTIF(\'MD Schools\'!$G$5:$G$200,"Safety")'
    summary["B11"] = '=COUNTIF(\'MD Schools\'!$G$5:$G$200,"Target")'
    summary["B12"] = '=COUNTIF(\'MD Schools\'!$G$5:$G$200,"Reach")'
    summary["D10"] = "Safety"
    summary["D11"] = "Target"
    summary["D12"] = "Reach"
    summary["E10"] = '=AVERAGEIF(\'MD Schools\'!$G$5:$G$200,"Safety",\'MD Schools\'!$E$5:$E$200)'
    summary["E11"] = '=AVERAGEIF(\'MD Schools\'!$G$5:$G$200,"Target",\'MD Schools\'!$E$5:$E$200)'
    summary["E12"] = '=AVERAGEIF(\'MD Schools\'!$G$5:$G$200,"Reach",\'MD Schools\'!$E$5:$E$200)'
    summary["F10"] = '=AVERAGEIF(\'MD Schools\'!$G$5:$G$200,"Safety",\'MD Schools\'!$F$5:$F$200)'
    summary["F11"] = '=AVERAGEIF(\'MD Schools\'!$G$5:$G$200,"Target",\'MD Schools\'!$F$5:$F$200)'
    summary["F12"] = '=AVERAGEIF(\'MD Schools\'!$G$5:$G$200,"Reach",\'MD Schools\'!$F$5:$F$200)'

    wb.save(WORKBOOK_PATH)


if __name__ == "__main__":
    main()

#!/usr/bin/env python3
"""Build an Excel checklist for school-specific application materials."""

from __future__ import annotations

import json
import re
import ssl
import urllib.request
from datetime import date
from pathlib import Path

import openpyxl
from bs4 import BeautifulSoup, Tag
from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation


ROOT = Path(__file__).resolve().parents[1]
INPUT_WORKBOOK = ROOT / "med_school_list_md.xlsx"
MSAR_SUMMARY = ROOT / "data" / "msar" / "2027" / "md_profiles_summary.json"
MSAR_PROFILES = ROOT / "data" / "msar" / "2027" / "profiles"
OUTPUT_WORKBOOK = ROOT / "med_school_application_materials_checklist.xlsx"

PROMPT_SOURCE_URL = (
    "https://www.shemmassianconsulting.com/blog/medical-school-secondary-essay-prompts"
)

PROMPT_ALIASES = {
    "Chicago Medical School at Rosalind Franklin University of Medicine & Science": "Chicago Medical School at Rosalind Franklin University of Medicine and Science",
    "Rush Medical College of Rush University Medical Center": "Rush Medical College of Rush University",
    "Robert Larner, M.D., College of Medicine at the University of Vermont": "University of Vermont Larner College of Medicine",
    "University of California, Davis, School of Medicine": "University of California – Davis School of Medicine",
    "California University of Science and Medicine-School of Medicine": "California University of Science and Medicine School of Medicine",
    "Lewis Katz School of Medicine at Temple University": "Temple University Lewis Katz School of Medicine",
    "NYU Grossman Long Island School of Medicine": "New York University Long Island School of Medicine",
    "University of Massachusetts T.H. Chan School of Medicine": "University of Massachusetts Medical School",
    "Wake Forest University School of Medicine": "Wake Forest School of Medicine",
    "Western Michigan University Homer Stryker M.D. School of Medicine": "Western Michigan University School of Medicine",
    "Rutgers, Robert Wood Johnson Medical School": "Rutgers Robert Wood Johnson Medical School",
    "Kaiser Permanente Bernard J. Tyson School of Medicine": "Kaiser Permanente School of Medicine",
    "University of California, Los Angeles David Geffen School of Medicine": "University of California – Los Angeles David Geffen School of Medicine",
    "East Tennessee State University James H. Quillen College of Medicine": "East Tennessee State University Quillen College of Medicine*",
}

STATUS_OPTIONS = ["Not started", "In progress", "Submitted", "Done", "N/A", "Blocked"]


def normalize(text: str | None) -> str:
    if not text:
        return ""
    return re.sub(r"\s+", " ", str(text)).strip()


def load_school_rows() -> list[dict]:
    wb = openpyxl.load_workbook(INPUT_WORKBOOK, data_only=True)
    ws = wb["MD Schools"]
    headers = [normalize(cell.value) for cell in ws[4]]
    rows = []
    for row in ws.iter_rows(min_row=5, values_only=True):
        if not row[0]:
            continue
        rows.append({headers[i]: row[i] for i in range(len(headers))})
    return rows


def load_msar_profiles() -> dict[str, dict]:
    summary = json.loads(MSAR_SUMMARY.read_text())
    profiles = {}
    for item in summary:
        profile_path = MSAR_PROFILES / f"{item['institutionId']}.json"
        profile = json.loads(profile_path.read_text())
        profile["_summary"] = item
        profiles[item["shortName"]] = profile
    return profiles


def is_school_heading(paragraph: Tag) -> bool:
    text = normalize(paragraph.get_text(" ", strip=True))
    if not text or text.startswith("(") or text.startswith("Time-sensitive"):
        return False
    if "Looking for secondary essay examples" in text:
        return False
    if not paragraph.find("a"):
        return False
    return paragraph.find("strong") is not None or len(text) < 140


def fetch_prompt_sections() -> dict[str, dict]:
    context = ssl._create_unverified_context()
    html = urllib.request.urlopen(PROMPT_SOURCE_URL, context=context, timeout=30).read()
    soup = BeautifulSoup(html, "html.parser")

    sections = {}
    for anchor in soup.find_all("a"):
        school_name = normalize(anchor.get_text(" ", strip=True))
        if not school_name or school_name.startswith("How to Get Into"):
            continue
        parent = anchor.find_parent("p")
        if parent is None or not is_school_heading(parent):
            continue

        notes: list[str] = []
        prompts: list[str] = []
        cycle = ""
        sibling = parent.next_sibling
        while sibling:
            if isinstance(sibling, Tag):
                if sibling.name in {"h2", "hr"}:
                    break
                if sibling.name == "p":
                    text = normalize(sibling.get_text(" ", strip=True))
                    if is_school_heading(sibling):
                        break
                    if not text or "Looking for secondary essay examples" in text:
                        break
                    if re.fullmatch(r"20\d{2}\s*[–—-]\s*20\d{2}", text):
                        cycle = text
                    elif text.startswith("(Suggested reading") or text.startswith("(Suggested Reading"):
                        pass
                    elif text.startswith("Time-sensitive"):
                        notes.append(text)
                    elif len(text) < 160 and (
                        "limit" in text.lower()
                        or "maximum" in text.lower()
                        or "all questions" in text.lower()
                    ):
                        notes.append(text)
                    else:
                        prompts.append(text)
                elif sibling.name in {"ol", "ul"}:
                    for item in sibling.find_all("li", recursive=False):
                        text = normalize(item.get_text(" ", strip=True))
                        if text:
                            prompts.append(text)
            sibling = sibling.next_sibling

        sections[school_name] = {"cycle": cycle, "notes": notes, "prompts": prompts}
    return sections


def classify_prompt(text: str) -> str:
    lower = text.lower()
    rules = [
        ("Why this school / mission fit", ["why", "reasons for applying", "interested in attending", "attracted you", "resonate"]),
        ("Mission, values, or institutional fit", ["mission", "values", "vision", "jesuit", "faith", "spiritual"]),
        ("Diversity, identity, or class contribution", ["diversity", "identity", "background", "unique", "perspective", "enrich", "contribute"]),
        ("Challenge, adversity, or resilience", ["challenge", "advers", "obstacle", "setback", "failure", "resilien", "difficult"]),
        ("Current or gap-year activities", ["current activities", "gap", "time off", "not currently enrolled", "plans for the", "full-time student"]),
        ("Reapplication update", ["reapplicant", "previously applied", "re-applicant", "prior application"]),
        ("Research or scholarly interests", ["research", "scholarly", "publication", "manuscript"]),
        ("Clinical exposure / motivation for medicine", ["clinical", "patient", "medicine", "physician", "exposure to medicine", "healthcare"]),
        ("Service, underserved communities, or health equity", ["service", "underserved", "equity", "social determinants", "disparit", "justice", "community"]),
        ("Teamwork, collaboration, or leadership", ["team", "collaborat", "leadership", "leader"]),
        ("Professionalism, integrity, or feedback", ["professionalism", "integrity", "feedback", "criticized", "humility", "ethic"]),
        ("Future goals or career path", ["future goals", "career", "ten years", "fifteen", "twenty years"]),
        ("Academic context or coursework", ["academic", "mcat", "gpa", "course", "grades", "standardized"]),
        ("Additional information", ["additional information", "anything else", "not reflected", "not included"]),
        ("Campus, program, pathway, or track interest", ["campus", "track", "pathway", "prime", "branch", "program"]),
        ("Portfolio artifact / competency demonstration", ["artifact", "portfolio", "show us", "compassion", "curiosity", "creativity"]),
        ("Family, alumni, state, or regional connection", ["family", "alumni", "state", "ties", "connection", "resident"]),
    ]
    for label, words in rules:
        if any(word in lower for word in words):
            return label
    return "School-specific short answer"


def prompt_limit(text: str, note_text: str = "") -> str:
    combined = f"{text} {note_text}"
    patterns = [
        r"\d{2,5}\s*[- ]?\s*(?:characters|character|words|word)",
        r"\d+\s*to\s*\d+\s*[- ]?\s*(?:characters|words)",
        r"no\s+(?:word|character)\s+(?:count|limit)",
    ]
    matches = []
    for pattern in patterns:
        matches.extend(re.findall(pattern, combined, flags=re.I))
    return "; ".join(dict.fromkeys(matches))


def required_status(text: str) -> str:
    lower = text.lower()
    if "optional" in lower:
        return "Optional/if applicable"
    if "reapplicant" in lower or "if applicable" in lower or "if yes" in lower:
        return "Conditional"
    return "Required"


def yn(value: str | None) -> str:
    return {"Y": "Yes", "N": "No", "C": "Conditional", "P": "Preferred", "A": "Accepted"}.get(
        normalize(value), normalize(value)
    )


def letter_summary(app: dict) -> str:
    if app.get("amcasLettersInd") == "N":
        return "AMCAS letters not required according to MSAR."
    min_letters = app.get("amcasLettersMin")
    max_letters = app.get("amcasLettersMax")
    counts = ""
    if min_letters and max_letters:
        counts = f"{min_letters} letters" if min_letters == max_letters else f"{min_letters}-{max_letters} letters"
    elif min_letters:
        counts = f"Minimum {min_letters} letters"
    elif max_letters:
        counts = f"Maximum {max_letters} letters"
    accepted = []
    if app.get("committeeLetterAccepted"):
        accepted.append(f"committee {yn(app.get('committeeLetterAccepted')).lower()}")
    if app.get("letterPacketsAccepted"):
        accepted.append(f"packet {yn(app.get('letterPacketsAccepted')).lower()}")
    if app.get("individualLetterAccepted"):
        accepted.append(f"individual {yn(app.get('individualLetterAccepted')).lower()}")
    details = normalize(app.get("letterGuidance") or app.get("amcasLettersText"))
    pieces = [counts, ", ".join(accepted), details]
    return " | ".join(piece for piece in pieces if piece)


def assessment_summary(app: dict) -> str:
    required = yn(app.get("additionalAssessmentsRequired"))
    info = normalize(app.get("additionalInfo"))
    keywords = []
    for word in ["Casper", "PREview", "Duet", "Kira", "situational judgment"]:
        if word.lower() in info.lower():
            keywords.append(word)
    if keywords:
        return f"{required}; {', '.join(dict.fromkeys(keywords))} noted. {info}"
    if required == "No":
        return "No additional assessment required per MSAR."
    return f"{required}. {info}" if info else required


def add_rows(ws, headers, rows):
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header, "") for header in headers])


def style_sheet(ws, freeze="A2"):
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = Border(bottom=thin)
    ws.freeze_panes = freeze
    ws.auto_filter.ref = ws.dimensions


def set_widths(ws, widths: dict[str, int]):
    for column, width in widths.items():
        ws.column_dimensions[column].width = width


def main() -> None:
    schools = load_school_rows()
    profiles = load_msar_profiles()
    prompt_sections = fetch_prompt_sections()

    wb = Workbook()
    wb.remove(wb.active)

    overview_rows = []
    checklist_rows = []
    essay_rows = []
    source_rows = []

    for school in schools:
        name = normalize(school["School"])
        profile = profiles[name]
        app = profile["medSchoolApplication"]
        summary = profile["_summary"]
        prompt_name = PROMPT_ALIASES.get(name, name)
        prompt_info = prompt_sections.get(prompt_name, {"cycle": "", "notes": [], "prompts": []})
        prompt_notes = "; ".join(dict.fromkeys(prompt_info["notes"]))
        prompts = prompt_info["prompts"]

        secondary_deadline = normalize(app.get("secApplLatestDtText"))
        overview_rows.append(
            {
                "School": name,
                "City": school.get("City"),
                "Reach/Target/Safety": school.get("Reach/Target/Safety"),
                "Reported GPA": school.get("Reported GPA"),
                "Reported MCAT": school.get("Reported MCAT"),
                "Secondary Required": yn(app.get("secApplReq")),
                "Secondary Deadline": secondary_deadline,
                "Secondary Fee": app.get("secFeeAmount") or "",
                "Fee Waiver": yn(app.get("secFeeWaiverInd")),
                "Letters": letter_summary(app),
                "Additional Assessment": assessment_summary(app),
                "MCAT Window": f"Oldest: {app.get('mcatOldestConsidered') or ''}; Latest: {app.get('mcatLatestConsidered') or ''}",
                "Updates/LOI": normalize(app.get("letterOfIntentAdditionalInfo")),
                "Technical Standards": app.get("technicalStandardLink") or normalize(app.get("technicalStandardInfo")),
                "Admissions URL": profile.get("admUrl") or "",
                "Prompt Cycle": prompt_info.get("cycle") or "Latest public listing found",
                "Prompt Source Notes": prompt_notes,
            }
        )

        common_items = [
            ("Primary application", "Submit AMCAS primary application", "AMCAS primary application service", app.get("priApplLatestDtText")),
            ("Secondary application", "Complete school secondary application", f"Fee: ${app.get('secFeeAmount') or 'see portal'}; waiver: {yn(app.get('secFeeWaiverInd'))}", secondary_deadline),
            ("Secondary essays", "Draft, revise, and proofread secondary essays", f"{len(prompts)} prompt rows in Essay Prompts sheet. {prompt_notes}", secondary_deadline),
            ("Letters of recommendation", "Assign letters in AMCAS", letter_summary(app), secondary_deadline),
            ("MCAT", "Confirm MCAT score received and within accepted date window", f"Oldest considered: {app.get('mcatOldestConsidered') or 'see MSAR'}; latest considered: {app.get('mcatLatestConsidered') or 'see MSAR'}", ""),
            ("Additional assessment", "Complete Casper/PREview/other assessment if required", assessment_summary(app), ""),
            ("Technical standards", "Review and be ready to attest to technical standards", app.get("technicalStandardLink") or normalize(app.get("technicalStandardInfo")), ""),
            ("Updates / letters of interest", "Record update-letter and letter-of-intent policy", normalize(app.get("letterOfIntentAdditionalInfo")), ""),
        ]
        for category, material, details, due in common_items:
            checklist_rows.append(
                {
                    "Done": "☐",
                    "Status": "Not started",
                    "School": name,
                    "Category": category,
                    "Required Material": material,
                    "Requirement Details": details,
                    "Due / Timing": due or "",
                    "Next Action": "",
                    "Source": "MSAR 2027 cache / prompt source",
                }
            )

        for index, prompt in enumerate(prompts, start=1):
            essay_rows.append(
                {
                    "Done": "☐",
                    "Status": "Not started",
                    "School": name,
                    "Prompt #": index,
                    "Required?": required_status(prompt),
                    "Topic Summary": classify_prompt(prompt),
                    "Limit": prompt_limit(prompt, prompt_notes),
                    "Exact Prompt Location": "Verify exact wording in school portal/source URL",
                    "Cycle Listed": prompt_info.get("cycle") or "",
                    "Source URL": PROMPT_SOURCE_URL,
                    "Draft File": "",
                    "Notes": prompt_notes,
                }
            )

        source_rows.append(
            {
                "School": name,
                "Admissions URL": profile.get("admUrl") or "",
                "Secondary URL": app.get("secApplUrl") or "",
                "Prompt Source URL": PROMPT_SOURCE_URL,
                "MSAR Profile Cache": str(MSAR_PROFILES / f"{summary['institutionId']}.json"),
                "MSAR Last Edited": profile.get("lastEditTime") or summary.get("lastEditTime") or "",
                "Notes": "Prompt topics are summarized to keep the checklist concise; verify exact wording in the live portal.",
            }
        )

    sheets = {
        "School Overview": (
            [
                "School",
                "City",
                "Reach/Target/Safety",
                "Reported GPA",
                "Reported MCAT",
                "Secondary Required",
                "Secondary Deadline",
                "Secondary Fee",
                "Fee Waiver",
                "Letters",
                "Additional Assessment",
                "MCAT Window",
                "Updates/LOI",
                "Technical Standards",
                "Admissions URL",
                "Prompt Cycle",
                "Prompt Source Notes",
            ],
            overview_rows,
        ),
        "Checklist": (
            [
                "Done",
                "Status",
                "School",
                "Category",
                "Required Material",
                "Requirement Details",
                "Due / Timing",
                "Next Action",
                "Source",
            ],
            checklist_rows,
        ),
        "Essay Prompts": (
            [
                "Done",
                "Status",
                "School",
                "Prompt #",
                "Required?",
                "Topic Summary",
                "Limit",
                "Exact Prompt Location",
                "Cycle Listed",
                "Source URL",
                "Draft File",
                "Notes",
            ],
            essay_rows,
        ),
        "Sources": (
            [
                "School",
                "Admissions URL",
                "Secondary URL",
                "Prompt Source URL",
                "MSAR Profile Cache",
                "MSAR Last Edited",
                "Notes",
            ],
            source_rows,
        ),
    }

    for title, (headers, rows) in sheets.items():
        ws = wb.create_sheet(title)
        add_rows(ws, headers, rows)
        style_sheet(ws)

    validations = [
        ("Checklist", "B"),
        ("Essay Prompts", "B"),
    ]
    for sheet_name, column in validations:
        ws = wb[sheet_name]
        validation = DataValidation(type="list", formula1=f'"{",".join(STATUS_OPTIONS)}"', allow_blank=False)
        ws.add_data_validation(validation)
        validation.add(f"{column}2:{column}{ws.max_row}")

    for sheet_name in ["Checklist", "Essay Prompts"]:
        ws = wb[sheet_name]
        green_fill = PatternFill("solid", fgColor="E2F0D9")
        ws.conditional_formatting.add(
            f"A2:I{ws.max_row}",
            FormulaRule(formula=["$B2=\"Done\""], fill=green_fill),
        )
        ws.conditional_formatting.add(
            f"A2:L{ws.max_row}",
            FormulaRule(formula=["$B2=\"Done\""], fill=green_fill),
        )

    set_widths(
        wb["School Overview"],
        {
            "A": 48,
            "B": 22,
            "C": 16,
            "D": 12,
            "E": 12,
            "F": 14,
            "G": 18,
            "H": 14,
            "I": 12,
            "J": 60,
            "K": 60,
            "L": 34,
            "M": 40,
            "N": 36,
            "O": 36,
            "P": 18,
            "Q": 48,
        },
    )
    set_widths(
        wb["Checklist"],
        {
            "A": 8,
            "B": 15,
            "C": 48,
            "D": 22,
            "E": 36,
            "F": 70,
            "G": 18,
            "H": 28,
            "I": 26,
        },
    )
    set_widths(
        wb["Essay Prompts"],
        {
            "A": 8,
            "B": 15,
            "C": 48,
            "D": 10,
            "E": 18,
            "F": 40,
            "G": 22,
            "H": 36,
            "I": 16,
            "J": 46,
            "K": 30,
            "L": 45,
        },
    )
    set_widths(
        wb["Sources"],
        {
            "A": 48,
            "B": 46,
            "C": 46,
            "D": 62,
            "E": 58,
            "F": 24,
            "G": 54,
        },
    )

    for ws in wb.worksheets:
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("http"):
                    cell.hyperlink = cell.value
                    cell.style = "Hyperlink"

    cover = wb.create_sheet("Read Me", 0)
    cover.append(["Med School Application Materials Checklist"])
    cover.append(["Generated", date.today().isoformat()])
    cover.append(["Source school list", str(INPUT_WORKBOOK)])
    cover.append(["Prompt source", PROMPT_SOURCE_URL])
    cover.append(["MSAR data", str(MSAR_SUMMARY)])
    cover.append(
        [
            "Note",
            "Use the Done box and Status dropdowns to track work. Essay rows summarize prompt topics; verify exact wording in each live secondary portal when received.",
        ]
    )
    style_sheet(cover)
    set_widths(cover, {"A": 22, "B": 110})

    wb.save(OUTPUT_WORKBOOK)
    print(f"Wrote {OUTPUT_WORKBOOK}")


if __name__ == "__main__":
    main()
